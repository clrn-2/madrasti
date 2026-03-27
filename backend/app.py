from fastapi import FastAPI, HTTPException, Header, Depends, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, EmailStr, ConfigDict
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any
from sqlalchemy import create_engine, select, and_, or_, text, func
from sqlalchemy.orm import sessionmaker, Session
from dotenv import load_dotenv
import jwt
import bcrypt
import re
from pydantic import TypeAdapter, ValidationError
from email.message import EmailMessage
import smtplib
import secrets
import hashlib
import hmac
import sys
import os
from collections import defaultdict
import time
import asyncio
from contextlib import asynccontextmanager

# Add parent directory to path for imports
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.insert(0, current_dir)
sys.path.insert(0, parent_dir)

load_dotenv(os.path.join(parent_dir, ".env"))

from models import (
    Base, School, Class, Student, User, AttendanceRecord, AttendanceSubmission,
    AuditLog, RoleEnum, AttendanceStatusEnum, SubmissionStatusEnum, PasswordResetRequest,
    SchoolTermSettings, Friendship, FriendshipStatusEnum, ChatSession, Message, UserBlock, ChatMute,
    CallSession, CallIceCandidate
)

# ===== Database Setup =====
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./madrasti.db")
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

# Create tables
Base.metadata.create_all(bind=engine)


def run_schema_migrations():
    """Small runtime migrations for SQLite local environments."""
    with engine.begin() as conn:
        user_columns = [row[1] for row in conn.execute(text("PRAGMA table_info(users)"))]
        if "phone" not in user_columns:
            conn.execute(text("ALTER TABLE users ADD COLUMN phone VARCHAR(50)"))
        if "specialization" not in user_columns:
            conn.execute(text("ALTER TABLE users ADD COLUMN specialization VARCHAR(255)"))
        if "profile_image" not in user_columns:
            conn.execute(text("ALTER TABLE users ADD COLUMN profile_image TEXT"))
        if "role_label" not in user_columns:
            conn.execute(text("ALTER TABLE users ADD COLUMN role_label VARCHAR(50)"))
        if "public_id" not in user_columns:
            conn.execute(text("ALTER TABLE users ADD COLUMN public_id VARCHAR(5)"))
        conn.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS idx_users_public_id_unique ON users(public_id)"))

        submission_columns = [row[1] for row in conn.execute(text("PRAGMA table_info(attendance_submissions)"))]
        if "academic_year" not in submission_columns:
            conn.execute(text(f"ALTER TABLE attendance_submissions ADD COLUMN academic_year INTEGER NOT NULL DEFAULT {datetime.utcnow().year}"))
        if "term" not in submission_columns:
            conn.execute(text("ALTER TABLE attendance_submissions ADD COLUMN term VARCHAR(20) NOT NULL DEFAULT 'first'"))
        if "deleted_at" not in submission_columns:
            conn.execute(text("ALTER TABLE attendance_submissions ADD COLUMN deleted_at DATETIME"))
        if "purge_at" not in submission_columns:
            conn.execute(text("ALTER TABLE attendance_submissions ADD COLUMN purge_at DATETIME"))

        # Ensure new tables created in older runtime DBs
        conn.execute(text(
            "CREATE TABLE IF NOT EXISTS user_blocks ("
            "id INTEGER PRIMARY KEY, "
            "blocker_id INTEGER NOT NULL, "
            "blocked_id INTEGER NOT NULL, "
            "created_at DATETIME"
            ")"
        ))
        conn.execute(text(
            "CREATE TABLE IF NOT EXISTS chat_mutes ("
            "id INTEGER PRIMARY KEY, "
            "user_id INTEGER NOT NULL, "
            "session_id INTEGER NOT NULL, "
            "is_muted BOOLEAN DEFAULT 0, "
            "created_at DATETIME, "
            "updated_at DATETIME"
            ")"
        ))
        conn.execute(text(
            "CREATE TABLE IF NOT EXISTS call_sessions ("
            "id INTEGER PRIMARY KEY, "
            "caller_id INTEGER NOT NULL, "
            "callee_id INTEGER NOT NULL, "
            "status VARCHAR(30) DEFAULT 'ringing', "
            "offer_sdp TEXT, "
            "answer_sdp TEXT, "
            "created_at DATETIME, "
            "updated_at DATETIME, "
            "ended_at DATETIME"
            ")"
        ))
        conn.execute(text(
            "CREATE TABLE IF NOT EXISTS call_ice_candidates ("
            "id INTEGER PRIMARY KEY, "
            "call_id INTEGER NOT NULL, "
            "sender_id INTEGER NOT NULL, "
            "recipient_id INTEGER NOT NULL, "
            "candidate TEXT NOT NULL, "
            "sdp_mid VARCHAR(255), "
            "sdp_mline_index INTEGER, "
            "created_at DATETIME"
            ")"
        ))

        # Add last_seen column to users if missing
        if "last_seen" not in user_columns:
            conn.execute(text("ALTER TABLE users ADD COLUMN last_seen DATETIME"))


run_schema_migrations()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ===== Configuration =====
SECRET_KEY = os.getenv("SECRET_KEY", "")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = int(os.getenv("ACCESS_TOKEN_EXPIRE_DAYS", "7"))
ARCHIVE_RETENTION_DAYS = int(os.getenv("ARCHIVE_RETENTION_DAYS", "3"))
MASTER_SCHOOL_CODE = os.getenv("MASTER_SCHOOL_CODE", "")
OTP_EXPIRE_MINUTES = 10
PASSWORD_RESET_EXPIRE_MINUTES = 15
ALLOW_DEV_OTP_FALLBACK = os.getenv("ALLOW_DEV_OTP_FALLBACK", "true").lower() == "true"
LAUNCH_SUPER_ADMIN_EMAIL = os.getenv("LAUNCH_SUPER_ADMIN_EMAIL", "")
LAUNCH_SUPER_ADMIN_PASSWORD = os.getenv("LAUNCH_SUPER_ADMIN_PASSWORD", "")
pending_email_verifications: Dict[str, Dict[str, Any]] = {}

# ===== Rate Limiting =====
_rate_store: Dict[str, List[float]] = defaultdict(list)

def _build_rate_store_key(request: Request, key: str, scope: Optional[str] = None) -> str:
    ip = request.client.host
    scope_value = str(scope or "").strip().lower()
    return f"{key}:{ip}:{scope_value}" if scope_value else f"{key}:{ip}"


def rate_limit(request: Request, key: str, max_calls: int, window: int, scope: Optional[str] = None):
    store_key = _build_rate_store_key(request, key, scope)
    now = time.time()
    _rate_store[store_key] = [t for t in _rate_store[store_key] if t > now - window]
    if len(_rate_store[store_key]) >= max_calls:
        raise HTTPException(status_code=429, detail="طلبات كثيرة، انتظر قليلاً وحاول مجدداً")
    _rate_store[store_key].append(now)


def clear_rate_limit(request: Request, key: str, scope: Optional[str] = None):
    store_key = _build_rate_store_key(request, key, scope)
    _rate_store.pop(store_key, None)


def generate_unique_public_id(db: Session) -> str:
    for _ in range(2000):
        candidate = str(secrets.randbelow(90000) + 10000)
        exists = db.query(User).filter(User.public_id == candidate).first()
        if not exists:
            return candidate
    raise HTTPException(status_code=500, detail="تعذر إنشاء رقم ID فريد")


def assign_missing_public_ids(db: Session):
    users_without_public_id = db.query(User).filter(
        or_(User.public_id.is_(None), User.public_id == "")
    ).all()
    for user in users_without_public_id:
        user.public_id = generate_unique_public_id(db)


def get_env_file_path() -> str:
    return os.path.join(parent_dir, ".env")


def get_smtp_config_issues() -> List[str]:
    load_dotenv(get_env_file_path(), override=True)

    required_keys = [
        "SMTP_HOST",
        "SMTP_PORT",
        "SMTP_USER",
        "SMTP_PASSWORD",
        "SMTP_FROM_EMAIL",
    ]

    placeholder_values = {
        "your_email@gmail.com",
        "your_app_password",
    }

    issues: List[str] = []
    for key in required_keys:
        value = (os.getenv(key) or "").strip()
        if not value:
            issues.append(f"{key}: missing")
            continue
        if value in placeholder_values:
            issues.append(f"{key}: placeholder")

    return issues

# ===== Request/Response Models =====

class UserBase(BaseModel):
    public_id: Optional[str] = None
    email: str
    full_name: str
    phone: Optional[str] = None
    specialization: Optional[str] = None
    profile_image: Optional[str] = None
    role: str = "teacher"
    role_label: Optional[str] = None
    school_name: Optional[str] = None

class UserCreate(UserBase):
    password: str
    school_id: int
    school_code: str
    verification_code: Optional[str] = None
    verification_token: Optional[str] = None


class RegisterVerificationRequest(BaseModel):
    email: str
    school_id: int
    school_code: str


class ProfileUpdateRequest(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    specialization: Optional[str] = None
    profile_image: Optional[str] = None


class ForgotPasswordRequest(BaseModel):
    identifier: str


class VerifyOtpRequest(BaseModel):
    identifier: str
    otp_code: str

class ConfirmPasswordResetRequest(BaseModel):
    token: Optional[str] = None
    identifier: Optional[str] = None
    otp_code: Optional[str] = None
    new_password: str
    confirm_password: str


class ChangePasswordRequest(BaseModel):
    current_password: str
    new_password: str
    confirm_password: str


class UserSearchItemResponse(BaseModel):
    id: int
    public_id: Optional[str] = None
    full_name: str
    email: str
    specialization: Optional[str] = None
    profile_image: Optional[str] = None
    role: str
    role_label: Optional[str] = None


class FriendRequestCreateRequest(BaseModel):
    receiver_id: int


class FriendRequestRespondRequest(BaseModel):
    action: str


class FriendRequestItemResponse(BaseModel):
    request_id: int
    user_id: int
    public_id: Optional[str] = None
    full_name: str
    specialization: Optional[str] = None
    profile_image: Optional[str] = None
    role_label: str
    status: str
    created_at: datetime


class BlockedUserItemResponse(BaseModel):
    user_id: int
    public_id: Optional[str] = None
    full_name: str
    specialization: Optional[str] = None
    profile_image: Optional[str] = None
    role_label: str
    blocked_at: datetime


class ChatMessageCreateRequest(BaseModel):
    content: str


class ChatMuteToggleRequest(BaseModel):
    is_muted: bool


class ChatMessageEditRequest(BaseModel):
    content: str


class UserProfileCardResponse(BaseModel):
    id: int
    public_id: Optional[str] = None
    full_name: str
    email: str
    phone: Optional[str] = None
    specialization: Optional[str] = None
    profile_image: Optional[str] = None
    school_name: Optional[str] = None
    role_label: Optional[str] = None


class AdminUserRoleUpdateRequest(BaseModel):
    role_label: str


class CallStartRequest(BaseModel):
    callee_id: int


class CallSdpRequest(BaseModel):
    sdp: str


class CallIceRequest(BaseModel):
    candidate: str
    sdp_mid: Optional[str] = None
    sdp_mline_index: Optional[int] = None


class CallEndRequest(BaseModel):
    status: Optional[str] = "ended"

class UserResponse(UserBase):
    id: int
    is_active: bool
    is_super_admin: bool = False
    created_at: datetime

    model_config = ConfigDict(from_attributes=True)

class LoginRequest(BaseModel):
    """نموذج طلب تسجيل الدخول"""
    email: str
    password: str
    school_id: Optional[int] = None
    school_code: Optional[str] = None  # Optional for super admin

class LoginResponse(BaseModel):
    """نموذج استجابة تسجيل الدخول"""
    token: str
    user: UserResponse
    message: str = "تم تسجيل الدخول بنجاح"


class FacebookLoginRequest(BaseModel):
    """Simulated Facebook login payload (access_token is optional in this local setup)"""
    access_token: Optional[str] = None
    email: EmailStr
    full_name: Optional[str] = None
    school_code: Optional[str] = None

class SchoolResponse(BaseModel):
    """نموذج استجابة المدرسة"""
    id: int
    name: str
    code: str
    status: str

    model_config = ConfigDict(from_attributes=True)

class SchoolsListResponse(BaseModel):
    """نموذج قائمة المدارس"""
    schools: List[SchoolResponse]
    total: int

    model_config = ConfigDict(from_attributes=True)


class SchoolCreateRequest(BaseModel):
    name: str
    code: str


class SchoolUpdateRequest(BaseModel):
    name: str
    code: str

class StudentResponse(BaseModel):
    """نموذج الطالب في الاستجابة"""
    id: int
    full_name: str

    model_config = ConfigDict(from_attributes=True)

class ClassResponse(BaseModel):
    id: int
    name: str
    school_id: int
    total_students: int = 0

    model_config = ConfigDict(from_attributes=True)

class ClassStudentsResponse(BaseModel):
    """نموذج استجابة طلاب الصف"""
    class_id: int
    class_name: str
    total_students: int
    students: List[StudentResponse]


class AdminClassUpsertRequest(BaseModel):
    name: str
    students: List[str]
    school_id: Optional[int] = None

class AttendanceRecordRequest(BaseModel):
    student_id: int
    status: str  # "present" or "absent"
    session_number: Optional[int] = None
    notes: Optional[str] = None

class AttendanceSubmissionRequest(BaseModel):
    class_id: int
    date: date
    submission_type: str  # "daily" or "per-session"
    num_sessions: Optional[int] = None
    records: List[AttendanceRecordRequest]

class AttendanceRecordResponse(BaseModel):
    id: int
    student_id: int
    date: date
    session_number: Optional[int]
    status: str
    notes: Optional[str]
    created_at: datetime

    model_config = ConfigDict(from_attributes=True)

class AttendanceSubmissionResponse(BaseModel):
    id: int
    class_id: int
    date: date
    daily_classes_count: int
    submission_type: str
    num_sessions: Optional[int]
    academic_year: int
    term: str
    status: str
    created_at: datetime
    submitted_at: Optional[datetime]
    records: List[AttendanceRecordResponse]

    model_config = ConfigDict(from_attributes=True)


class AttendanceSubmissionSummaryResponse(BaseModel):
    id: int
    class_id: int
    teacher_name: str
    class_name: str
    date: date
    academic_year: int
    term: str
    total_students: int
    present_count: int
    absent_count: int


class SchoolTermSettingsUpdateRequest(BaseModel):
    academic_year: int
    term: str


class AttendanceArchiveTermRequest(BaseModel):
    school_id: Optional[int] = None
    academic_year: int
    term: str


class SubmissionStudentDetailResponse(BaseModel):
    record_id: int
    student_id: int
    student_name: str
    status: str


class AttendanceSubmissionDetailResponse(BaseModel):
    submission_id: int
    teacher_name: str
    class_name: str
    date: date
    students: List[SubmissionStudentDetailResponse]


class StudentAttendanceDateResponse(BaseModel):
    date: date
    status: str


class StudentAttendanceReportItemResponse(BaseModel):
    student_id: int
    student_name: str
    class_name: str
    present_days: int
    absent_days: int
    details: List[StudentAttendanceDateResponse]


class AttendanceRecordUpdateItemRequest(BaseModel):
    record_id: int
    status: str
    notes: Optional[str] = None


class AttendanceSubmissionUpdateRequest(BaseModel):
    records: List[AttendanceRecordUpdateItemRequest]

class DailyReportResponse(BaseModel):
    """نموذج تقرير اليوم"""
    class_name: str
    date: date
    total_students: int
    present_count: int
    absent_count: int
    attendance_percentage: float

# ===== Authentication Utilities =====

def hash_password(password: str) -> str:
    """Hash password using bcrypt"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    """Verify password against hash"""
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))


def ensure_launch_super_admin():
    db = SessionLocal()
    try:
        default_school = db.query(School).order_by(School.id.asc()).first()
        if not default_school:
            default_school = School(name="مدرسة النظام المركزية", code="SYS-0000", status="active")
            db.add(default_school)
            db.flush()

        user = db.query(User).filter(User.email == LAUNCH_SUPER_ADMIN_EMAIL).first()
        if user:
            user.school_id = default_school.id
            user.full_name = user.full_name or "المسؤول الأعلى"
            user.password_hash = hash_password(LAUNCH_SUPER_ADMIN_PASSWORD)
            user.role = RoleEnum.ADMIN
            user.role_label = "super_admin"
            user.is_active = True
            user.is_super_admin = True
        else:
            user = User(
                school_id=default_school.id,
                public_id=generate_unique_public_id(db),
                email=LAUNCH_SUPER_ADMIN_EMAIL,
                full_name="المسؤول الأعلى",
                password_hash=hash_password(LAUNCH_SUPER_ADMIN_PASSWORD),
                role=RoleEnum.ADMIN,
                role_label="super_admin",
                is_active=True,
                is_super_admin=True,
            )
            db.add(user)

        assign_missing_public_ids(db)

        db.commit()
    finally:
        db.close()


ensure_launch_super_admin()


def hash_secret(value: str) -> str:
    return hashlib.sha256((value or "").encode("utf-8")).hexdigest()


def normalize_phone(raw_phone: str) -> str:
    cleaned = "".join(ch for ch in str(raw_phone or "") if ch.isdigit())
    if not cleaned:
        return ""
    if cleaned.startswith("962") and len(cleaned) >= 12:
        return cleaned
    if cleaned.startswith("0"):
        return f"962{cleaned[1:]}"
    if cleaned.startswith("7"):
        return f"962{cleaned}"
    return cleaned


def validate_password_policy_or_raise(password: str):
    pwd = password or ""
    if (
        len(pwd) < 8
        or not re.search(r"[A-Z]", pwd)
        or not re.search(r"[a-z]", pwd)
        or not re.search(r"[0-9]", pwd)
    ):
        raise HTTPException(
            status_code=400,
            detail="كلمة المرور يجب أن تكون 8 أحرف على الأقل وتحتوي على حرف كبير وحرف صغير ورقم"
        )


def find_user_by_identifier(db: Session, identifier: str) -> Optional[User]:
    normalized_identifier = (identifier or "").strip()
    if not normalized_identifier:
        return None

    if "@" in normalized_identifier:
        normalized_email = validate_email_or_raise(normalized_identifier)
        return db.query(User).filter(User.email == normalized_email).first()

    normalized_phone = normalize_phone(normalized_identifier)
    if not normalized_phone:
        return None
    return db.query(User).filter(User.phone == normalized_phone).first()


def validate_email_or_raise(email: str) -> str:
    normalized_email = (email or "").strip().lower()
    if not normalized_email:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني غير صالح")
    try:
        TypeAdapter(EmailStr).validate_python(normalized_email)
    except Exception:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني غير صالح")
    return normalized_email


def get_school_by_code_or_raise(db: Session, school_code: str) -> School:
    normalized_code = (school_code or "").strip()
    school = db.query(School).filter(School.code == normalized_code).first()
    if not school:
        raise HTTPException(status_code=400, detail="كود المدرسة غير صحيح")
    return school


def ensure_school_code_matches_selection_or_raise(db: Session, school: School, selected_school_id: int):
    selected_school = db.query(School).filter(School.id == selected_school_id).first()
    if not selected_school:
        raise HTTPException(status_code=400, detail="المدرسة المختارة غير موجودة")
    if school.id != selected_school_id:
        raise HTTPException(status_code=400, detail="كود المدرسة لا يطابق المدرسة المختارة")


def send_otp_email(email: str, code: str):
    load_dotenv(get_env_file_path(), override=True)

    smtp_host = os.getenv("SMTP_HOST")
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_use_tls = os.getenv("SMTP_USE_TLS", "true").lower() == "true"
    from_email = os.getenv("SMTP_FROM_EMAIL", smtp_user or "")

    issues = get_smtp_config_issues()
    if issues:
        detail = "إعدادات البريد الإلكتروني غير مكتملة: " + ", ".join(issues)
        raise HTTPException(status_code=500, detail=detail)

    msg = EmailMessage()
    msg["Subject"] = "كود التحقق - Madrasti"
    msg["From"] = from_email
    msg["To"] = email
    msg.set_content(
        f"كود التحقق الخاص بك هو: {code}\n"
        f"صلاحية الكود: {OTP_EXPIRE_MINUTES} دقائق."
    )

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
            if smtp_use_tls:
                server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
    except Exception as exc:
        raise HTTPException(status_code=500, detail="تعذر إرسال كود التحقق إلى البريد الإلكتروني") from exc


def verify_registration_otp_or_raise(request: UserCreate, school: School):
    verification_token = (request.verification_token or "").strip()
    entered_code = (request.verification_code or "").strip()

    if not verification_token or not entered_code:
        raise HTTPException(status_code=400, detail="كود التحقق غير صحيح")

    pending_record = pending_email_verifications.get(verification_token)
    if not pending_record:
        raise HTTPException(status_code=400, detail="كود التحقق غير صحيح")

    now = datetime.utcnow()
    if pending_record["expires_at"] < now:
        pending_email_verifications.pop(verification_token, None)
        raise HTTPException(status_code=400, detail="انتهت صلاحية كود التحقق")

    if (
        pending_record["email"] != request.email.lower()
        or pending_record["school_id"] != school.id
        or pending_record["code"] != entered_code
    ):
        raise HTTPException(status_code=400, detail="كود التحقق غير صحيح")

    pending_email_verifications.pop(verification_token, None)


def parse_binary_attendance_status(status: str) -> AttendanceStatusEnum:
    normalized = (status or '').strip().lower()
    if normalized == AttendanceStatusEnum.PRESENT.value:
        return AttendanceStatusEnum.PRESENT
    if normalized == AttendanceStatusEnum.ABSENT.value:
        return AttendanceStatusEnum.ABSENT
    raise HTTPException(status_code=400, detail="الحالات المدعومة فقط: حاضر أو غائب")

def create_access_token(user_id: int) -> str:
    """Create JWT token"""
    payload = {
        "sub": str(user_id),
        "exp": datetime.utcnow() + timedelta(days=ACCESS_TOKEN_EXPIRE_DAYS)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

def verify_token(token: str) -> Optional[int]:
    """Verify JWT token and return user_id"""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = int(payload.get("sub"))
        return user_id
    except (jwt.InvalidTokenError, ValueError):
        return None

async def get_current_user(
    authorization: Optional[str] = Header(None),
    db: Session = Depends(get_db)
) -> User:
    """Get current authenticated user"""
    if not authorization:
        raise HTTPException(status_code=401, detail="رمز الوصول مفقود")
    
    try:
        if authorization.startswith("Bearer "):
            token = authorization[7:]
        else:
            token = authorization
        
        user_id = verify_token(token)
        if not user_id:
            raise HTTPException(status_code=401, detail="رمز الوصول غير صحيح")
        
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="المستخدم غير موجود")
        
        return user
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=401, detail="خطأ في التحقق: " + str(e))


def ensure_admin_user(current_user: User):
    if not current_user.is_super_admin and current_user.role != RoleEnum.ADMIN:
        raise HTTPException(status_code=403, detail="هذه العملية متاحة للإدارة فقط")


def ensure_super_admin_user(current_user: User):
    if not current_user.is_super_admin:
        raise HTTPException(status_code=403, detail="هذه العملية متاحة لمسؤول النظام فقط")


def get_account_type(current_user: User) -> str:
    role_label = (current_user.role_label or "").strip().lower()
    if role_label in ("teacher", "admin", "principal"):
        return role_label
    if current_user.role == RoleEnum.ADMIN:
        return "admin"
    return "teacher"


def ensure_principal_user(current_user: User):
    if current_user.is_super_admin:
        return
    if get_account_type(current_user) != "principal":
        raise HTTPException(status_code=403, detail="هذه العملية متاحة لمدير المدرسة فقط")


def normalize_term_or_raise(term: str) -> str:
    value = (term or "").strip().lower()
    mapping = {
        "first": "first",
        "second": "second",
        "third": "third",
        "الفصل الاول": "first",
        "الفصل الأول": "first",
        "الاول": "first",
        "الأول": "first",
        "الفصل الثاني": "second",
        "الثاني": "second",
        "الفصل الثالث": "third",
        "الثالث": "third",
    }
    normalized = mapping.get(value, value)
    if normalized not in ("first", "second", "third"):
        raise HTTPException(status_code=400, detail="الفصل يجب أن يكون: first أو second أو third")
    return normalized


def ensure_school_term_settings(db: Session, school_id: int) -> SchoolTermSettings:
    settings = db.query(SchoolTermSettings).filter(SchoolTermSettings.school_id == school_id).first()
    if settings:
        return settings

    settings = SchoolTermSettings(
        school_id=school_id,
        current_academic_year=datetime.utcnow().year,
        current_term="first"
    )
    db.add(settings)
    db.commit()
    db.refresh(settings)
    return settings


def purge_expired_archived_submissions(db: Session):
    now = datetime.utcnow()
    expired = db.query(AttendanceSubmission).filter(
        AttendanceSubmission.deleted_at.isnot(None),
        AttendanceSubmission.purge_at.isnot(None),
        AttendanceSubmission.purge_at <= now
    ).all()
    if not expired:
        return

    for sub in expired:
        db.delete(sub)
    db.commit()


def resolve_target_school_id(request_school_id: Optional[int], current_user: User) -> int:
    if current_user.is_super_admin:
        return request_school_id or current_user.school_id

    if request_school_id and request_school_id != current_user.school_id:
        raise HTTPException(status_code=403, detail="غير مصرح لإدارة مدرسة أخرى")

    return current_user.school_id


def normalize_student_names(student_names: List[str]) -> List[str]:
    names = [n.strip() for n in student_names if isinstance(n, str) and n.strip()]
    if not names:
        raise HTTPException(status_code=400, detail="قائمة الطلاب مطلوبة")
    return names


def sync_students_for_class(db: Session, class_obj: Class, student_names: List[str]):
    names = normalize_student_names(student_names)
    existing_students = db.query(Student).filter(Student.class_id == class_obj.id).order_by(Student.id).all()

    for index, name in enumerate(names):
        if index < len(existing_students):
            existing_students[index].full_name = name
            existing_students[index].is_active = True
        else:
            db.add(Student(class_id=class_obj.id, full_name=name, is_active=True))

    for index in range(len(names), len(existing_students)):
        existing_students[index].is_active = False


def get_or_create_chat_session(db: Session, user_a_id: int, user_b_id: int) -> ChatSession:
    session_obj = db.query(ChatSession).filter(
        or_(
            and_(ChatSession.starter_id == user_a_id, ChatSession.joiner_id == user_b_id),
            and_(ChatSession.starter_id == user_b_id, ChatSession.joiner_id == user_a_id),
        )
    ).first()
    if session_obj:
        return session_obj

    session_obj = ChatSession(starter_id=user_a_id, joiner_id=user_b_id)
    db.add(session_obj)
    db.flush()
    return session_obj


def get_friendship_between(db: Session, user_a_id: int, user_b_id: int) -> Optional[Friendship]:
    return db.query(Friendship).filter(
        or_(
            and_(Friendship.sender_id == user_a_id, Friendship.receiver_id == user_b_id),
            and_(Friendship.sender_id == user_b_id, Friendship.receiver_id == user_a_id),
        )
    ).first()


def is_blocked_between(db: Session, user_a_id: int, user_b_id: int) -> bool:
    record = db.query(UserBlock).filter(
        or_(
            and_(UserBlock.blocker_id == user_a_id, UserBlock.blocked_id == user_b_id),
            and_(UserBlock.blocker_id == user_b_id, UserBlock.blocked_id == user_a_id),
        )
    ).first()
    return record is not None

# ===== Background Scheduler =====

async def _periodic_purge_loop():
    """Runs every hour and permanently deletes submissions whose purge_at has passed."""
    while True:
        await asyncio.sleep(3600)  # wait 1 hour between runs
        db = SessionLocal()
        try:
            purge_expired_archived_submissions(db)
        except Exception:
            pass
        finally:
            db.close()


@asynccontextmanager
async def lifespan(app_instance):
    task = asyncio.create_task(_periodic_purge_loop())
    yield
    task.cancel()
    try:
        await task
    except asyncio.CancelledError:
        pass


# ===== Create FastAPI App =====

app = FastAPI(
    lifespan=lifespan,
    title="Madrasti - نظام الحضور والغياب",
    description="تطبيق متكامل لإدارة الحضور والغياب للمدارس",
    version="1.0.0"
)

# Add CORS middleware - Allow requests from frontend
_cors_origins = [
    "http://localhost:8001",
    "http://127.0.0.1:8001",
    "http://localhost:3000",
    "https://localhost:3000",
    "http://localhost:5500",
    "http://127.0.0.1:5500",
]
_extra_origin = os.getenv("CORS_ORIGIN", "")
if _extra_origin:
    _cors_origins.append(_extra_origin)

app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_origin_regex=".*",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ===== Root Routes =====

@app.get("/health")
async def health_check(db: Session = Depends(get_db)):
    """Health check endpoint"""
    try:
        # Try to query a school to verify DB connection
        school = db.query(School).first()
        return {
            "status": "healthy",
            "timestamp": datetime.utcnow().isoformat(),
            "database": "connected",
            "schools_count": db.query(School).count()
        }
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Database error: {str(e)}")

# ===== Authentication Routes =====

@app.post("/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest, req: Request, db: Session = Depends(get_db)):
    normalized_email = validate_email_or_raise(request.email)
    rate_limit(req, "login", max_calls=10, window=60, scope=normalized_email)
    """تسجيل الدخول"""
    request.email = normalized_email
    
    # Find user first by email
    user = db.query(User).filter(User.email == request.email).first()

    if not user:
        raise HTTPException(status_code=401, detail="البريد الإلكتروني غير صحيح")

    if not verify_password(request.password, user.password_hash):
        raise HTTPException(status_code=401, detail="كلمة السر غير صحيحة")

    if not user.public_id:
        user.public_id = generate_unique_public_id(db)
        db.commit()
        db.refresh(user)
    
    if not user.is_active:
        raise HTTPException(status_code=401, detail="الحساب غير فعال")
    
    entered_school_code = (request.school_code or "").strip()
    selected_school_id = request.school_id

    # Master code is server-side only and reserved for super admin accounts
    if MASTER_SCHOOL_CODE and entered_school_code == MASTER_SCHOOL_CODE and not user.is_super_admin:
        raise HTTPException(status_code=403, detail="رمز المالك مخصص لمسؤول النظام فقط")

    if user.is_super_admin:
        target_school = None
        if selected_school_id:
            target_school = db.query(School).filter(School.id == selected_school_id).first()
            if not target_school:
                raise HTTPException(status_code=400, detail="المدرسة المختارة غير موجودة")
        else:
            target_school = db.query(School).filter(School.id == user.school_id).first()
            if not target_school:
                target_school = db.query(School).order_by(School.id.asc()).first()

        if target_school and user.school_id != target_school.id:
            user.school_id = target_school.id
            db.commit()
            db.refresh(user)
    else:
        # Login must always be tied to the selected school + its code for regular users
        if not selected_school_id:
            raise HTTPException(status_code=400, detail="يرجى اختيار المدرسة")
        if not entered_school_code:
            raise HTTPException(status_code=400, detail="رمز المدرسة مطلوب")

        selected_school = db.query(School).filter(School.id == selected_school_id).first()
        if not selected_school:
            raise HTTPException(status_code=400, detail="المدرسة المختارة غير موجودة")

        entered_school = db.query(School).filter(School.code == entered_school_code).first()
        if not entered_school:
            raise HTTPException(status_code=401, detail="كود المدرسة غير صحيح")

        if selected_school.code != entered_school_code:
            raise HTTPException(status_code=401, detail="كود المدرسة لا يطابق المدرسة المختارة")

        if user.school_id != selected_school.id:
            user.school_id = selected_school.id
            db.commit()
            db.refresh(user)
    
    # Create token
    token = create_access_token(user.id)
    clear_rate_limit(req, "login", scope=normalized_email)
    
    return LoginResponse(
        token=token,
        user=UserResponse(
            id=user.id,
            public_id=user.public_id,
            email=user.email,
            full_name=user.full_name,
            phone=user.phone,
            specialization=user.specialization,
            profile_image=user.profile_image,
            role=user.role.value,
            role_label=user.role_label,
            school_name=user.school.name if user.school else None,
            is_active=user.is_active,
            is_super_admin=user.is_super_admin,
            created_at=user.created_at
        ),
        message="تم تسجيل الدخول بنجاح"
    )

@app.post("/auth/logout")
async def logout(current_user: User = Depends(get_current_user)):
    """تسجيل الخروج"""
    return {"message": "تم تسجيل الخروج بنجاح"}


@app.put("/auth/profile", response_model=UserResponse)
async def update_profile(
    request: ProfileUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """تحديث ملف المستخدم (الاسم/الهاتف/التخصص) مع حفظ دائم في قاعدة البيانات."""
    if request.full_name is not None:
        normalized_name = " ".join(str(request.full_name).strip().split())
        if len(normalized_name) < 2:
            raise HTTPException(status_code=400, detail="الاسم الكامل غير صالح")
        current_user.full_name = normalized_name

    if request.specialization is not None:
        normalized_specialization = " ".join(str(request.specialization).strip().split())
        current_user.specialization = normalized_specialization or None

    if request.profile_image is not None:
        raw_profile_image = str(request.profile_image or "").strip()
        if not raw_profile_image:
            current_user.profile_image = None
        elif raw_profile_image.startswith("data:image/") and ";base64," in raw_profile_image:
            current_user.profile_image = raw_profile_image
        else:
            raise HTTPException(status_code=400, detail="صيغة صورة الملف الشخصي غير مدعومة")

    if request.phone is not None:
        normalized_phone = normalize_phone(request.phone)
        if normalized_phone:
            existing_phone_owner = db.query(User).filter(
                User.phone == normalized_phone,
                User.id != current_user.id
            ).first()
            if existing_phone_owner:
                raise HTTPException(status_code=400, detail="رقم الهاتف مستخدم من حساب آخر")
            current_user.phone = normalized_phone
        else:
            current_user.phone = None

    db.commit()
    db.refresh(current_user)

    return UserResponse(
        id=current_user.id,
        public_id=current_user.public_id,
        email=current_user.email,
        full_name=current_user.full_name,
        phone=current_user.phone,
        specialization=current_user.specialization,
        profile_image=current_user.profile_image,
        role=current_user.role.value,
        role_label=current_user.role_label,
        school_name=current_user.school.name if current_user.school else None,
        is_active=current_user.is_active,
        is_super_admin=current_user.is_super_admin,
        created_at=current_user.created_at
    )


@app.post("/auth/password/change")
async def change_password(
    request: ChangePasswordRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    current_password = request.current_password or ""
    new_password = request.new_password or ""
    confirm_password = request.confirm_password or ""

    if not verify_password(current_password, current_user.password_hash):
        raise HTTPException(status_code=400, detail="كلمة المرور الحالية غير صحيحة")

    if new_password != confirm_password:
        raise HTTPException(status_code=400, detail="تأكيد كلمة المرور غير مطابق")

    if current_password == new_password:
        raise HTTPException(status_code=400, detail="كلمة المرور الجديدة يجب أن تختلف عن الحالية")

    validate_password_policy_or_raise(new_password)
    current_user.password_hash = hash_password(new_password)
    current_user.updated_at = datetime.utcnow()
    db.commit()

    return {"message": "تم تغيير كلمة السر بنجاح"}


@app.post("/auth/password/forgot/request")
async def request_password_reset(request: ForgotPasswordRequest, db: Session = Depends(get_db)):
    """Start forgot-password flow with email link or phone OTP."""
    identifier = (request.identifier or "").strip()
    if not identifier:
        raise HTTPException(status_code=400, detail="أدخل البريد الإلكتروني أو رقم الهاتف")

    user = find_user_by_identifier(db, identifier)
    if not user:
        raise HTTPException(status_code=404, detail="البريد الإلكتروني أو رقم الهاتف غير مسجل")

    channel = "email" if "@" in identifier else "phone"
    expires_at = datetime.utcnow() + timedelta(minutes=PASSWORD_RESET_EXPIRE_MINUTES)

    token_raw = secrets.token_urlsafe(32)
    otp_raw = f"{secrets.randbelow(900000) + 100000}"

    old_requests = db.query(PasswordResetRequest).filter(
        PasswordResetRequest.user_id == user.id,
        PasswordResetRequest.used_at.is_(None)
    ).all()
    for old_req in old_requests:
        old_req.used_at = datetime.utcnow()

    reset_req = PasswordResetRequest(
        user_id=user.id,
        channel=channel,
        token_hash=hash_secret(token_raw),
        otp_hash=hash_secret(otp_raw),
        expires_at=expires_at
    )
    db.add(reset_req)
    db.commit()

    frontend_reset_base = os.getenv("FRONTEND_RESET_URL", "http://127.0.0.1:8001/reset_password.html")
    reset_link = f"{frontend_reset_base}?token={token_raw}"

    response = {
        "message": "تم إرسال رابط/كود إعادة التعيين، يرجى التحقق من بريدك أو هاتفك",
        "channel": channel,
        "expires_in_seconds": PASSWORD_RESET_EXPIRE_MINUTES * 60,
    }

    if channel == "email":
        try:
            load_dotenv(get_env_file_path(), override=True)
            msg = EmailMessage()
            msg["Subject"] = "إعادة تعيين كلمة المرور - Madrasti"
            msg["From"] = os.getenv("SMTP_FROM_EMAIL", os.getenv("SMTP_USER", ""))
            msg["To"] = user.email
            msg.set_content(
                "طلبت إعادة تعيين كلمة المرور.\n"
                f"رابط إعادة التعيين: {reset_link}\n"
                f"ينتهي الرابط خلال {PASSWORD_RESET_EXPIRE_MINUTES} دقيقة.\n"
                f"(كود بديل: {otp_raw})"
            )

            smtp_host = os.getenv("SMTP_HOST")
            smtp_user = os.getenv("SMTP_USER")
            smtp_password = os.getenv("SMTP_PASSWORD")
            smtp_port = int(os.getenv("SMTP_PORT", "587"))
            smtp_use_tls = os.getenv("SMTP_USE_TLS", "true").lower() == "true"

            with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
                if smtp_use_tls:
                    server.starttls()
                server.login(smtp_user, smtp_password)
                server.send_message(msg)
        except Exception:
            if not ALLOW_DEV_OTP_FALLBACK:
                raise HTTPException(status_code=500, detail="تعذر إرسال رابط إعادة التعيين")
            print(f"[DEV RESET LINK] {user.email} => {reset_link}")
            print(f"[DEV RESET OTP] {user.email} => {otp_raw}")
            response["dev_reset_link"] = reset_link
            response["dev_otp_code"] = otp_raw
    else:
        # SMS provider is not configured in this local project; dev fallback for OTP.
        print(f"[DEV SMS OTP] {user.phone} => {otp_raw}")
        if ALLOW_DEV_OTP_FALLBACK:
            response["dev_otp_code"] = otp_raw

    return response


@app.post("/auth/password/forgot/verify-otp")
async def verify_reset_otp(request: VerifyOtpRequest, db: Session = Depends(get_db)):
    """Verify OTP code is correct and not expired, without consuming it."""
    identifier = (request.identifier or "").strip()
    otp_code_raw = (request.otp_code or "").strip()

    if not identifier or not otp_code_raw:
        raise HTTPException(status_code=400, detail="أدخل البريد/الهاتف وكود التحقق")

    user = find_user_by_identifier(db, identifier)
    if not user:
        raise HTTPException(status_code=404, detail="البريد الإلكتروني أو رقم الهاتف غير مسجل")

    otp_hash = hash_secret(otp_code_raw)
    reset_request = db.query(PasswordResetRequest).filter(
        PasswordResetRequest.user_id == user.id,
        PasswordResetRequest.otp_hash == otp_hash,
        PasswordResetRequest.used_at.is_(None)
    ).order_by(PasswordResetRequest.id.desc()).first()

    if not reset_request:
        raise HTTPException(status_code=400, detail="كود التحقق غير صحيح")

    if reset_request.expires_at < datetime.utcnow():
        raise HTTPException(status_code=400, detail="انتهت صلاحية كود التحقق")

    return {"message": "كود التحقق صحيح"}


@app.post("/auth/password/forgot/confirm")
async def confirm_password_reset(request: ConfirmPasswordResetRequest, db: Session = Depends(get_db)):
    """Confirm reset via token link or identifier + OTP and set new password."""
    new_password = request.new_password or ""
    confirm_password = request.confirm_password or ""

    if new_password != confirm_password:
        raise HTTPException(status_code=400, detail="تأكيد كلمة المرور غير مطابق")

    validate_password_policy_or_raise(new_password)

    reset_request = None
    user = None

    if request.token:
        token_hash = hash_secret(request.token.strip())
        reset_request = db.query(PasswordResetRequest).filter(
            PasswordResetRequest.token_hash == token_hash,
            PasswordResetRequest.used_at.is_(None)
        ).order_by(PasswordResetRequest.id.desc()).first()
    else:
        identifier = (request.identifier or "").strip()
        otp_code = (request.otp_code or "").strip()
        if not identifier or not otp_code:
            raise HTTPException(status_code=400, detail="أدخل الرابط أو (البريد/الهاتف + كود التحقق)")

        user = find_user_by_identifier(db, identifier)
        if not user:
            raise HTTPException(status_code=404, detail="البريد الإلكتروني أو رقم الهاتف غير مسجل")

        otp_hash = hash_secret(otp_code)
        reset_request = db.query(PasswordResetRequest).filter(
            PasswordResetRequest.user_id == user.id,
            PasswordResetRequest.otp_hash == otp_hash,
            PasswordResetRequest.used_at.is_(None)
        ).order_by(PasswordResetRequest.id.desc()).first()

    if not reset_request:
        raise HTTPException(status_code=400, detail="رابط/كود إعادة التعيين غير صحيح")

    if reset_request.expires_at < datetime.utcnow():
        raise HTTPException(status_code=400, detail="انتهت صلاحية رابط/كود إعادة التعيين")

    if not user:
        user = db.query(User).filter(User.id == reset_request.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")

    user.password_hash = hash_password(new_password)
    reset_request.used_at = datetime.utcnow()
    db.commit()

    return {"message": "تم تغيير كلمة السر بنجاح"}


@app.post("/auth/register/send-verification")
async def send_register_verification(request: RegisterVerificationRequest, db: Session = Depends(get_db)):
    """إرسال كود OTP للتسجيل بعد التحقق من كود المدرسة وصيغة البريد"""

    email = validate_email_or_raise(request.email)
    school = get_school_by_code_or_raise(db, request.school_code)
    ensure_school_code_matches_selection_or_raise(db, school, request.school_id)

    existing_user = db.query(User).filter(User.email == email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني مسجل بالفعل")

    verification_code = f"{secrets.randbelow(900000) + 100000}"
    verification_token = secrets.token_urlsafe(32)
    expires_at = datetime.utcnow() + timedelta(minutes=OTP_EXPIRE_MINUTES)

    delivery_method = "smtp"
    dev_verification_code: Optional[str] = None

    try:
        send_otp_email(email, verification_code)
    except HTTPException as exc:
        if exc.status_code == 500 and ALLOW_DEV_OTP_FALLBACK:
            delivery_method = "dev_fallback"
            dev_verification_code = verification_code
            print(f"[DEV OTP] {email} => {verification_code}")
        else:
            raise

    pending_email_verifications[verification_token] = {
        "email": email,
        "school_id": school.id,
        "code": verification_code,
        "expires_at": expires_at,
    }

    response = {
        "message": "تم إرسال كود التحقق إلى البريد الإلكتروني",
        "verification_token": verification_token,
        "expires_in_seconds": OTP_EXPIRE_MINUTES * 60,
        "delivery_method": delivery_method,
    }

    if dev_verification_code:
        response["verification_code"] = dev_verification_code
        response["message"] = "تم إنشاء كود تحقق للاختبار المحلي (SMTP غير مهيأ)"

    return response


@app.get("/auth/register/smtp-status")
async def smtp_status():
    issues = get_smtp_config_issues()
    return {
        "ok": len(issues) == 0,
        "allow_dev_otp_fallback": ALLOW_DEV_OTP_FALLBACK,
        "issues": issues,
        "message": "جاهز" if len(issues) == 0 else "إعدادات SMTP تحتاج تعديل"
    }

@app.post("/auth/register", response_model=UserResponse)
async def register(request: UserCreate, db: Session = Depends(get_db)):
    """تسجيل مستخدم جديد (للمعلمين فقط)"""

    request.email = validate_email_or_raise(request.email)
    validate_password_policy_or_raise(request.password)
    school = get_school_by_code_or_raise(db, request.school_code)
    ensure_school_code_matches_selection_or_raise(db, school, request.school_id)
    
    # Check if email already exists
    existing_user = db.query(User).filter(User.email == request.email).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني مسجل بالفعل")
    
    # Create new user based on role
    requested_role = (request.role or "teacher").lower()
    is_super_admin_role = requested_role == "super_admin"
    if is_super_admin_role:
        raise HTTPException(status_code=403, detail="لا يمكن إنشاء حساب مسؤول النظام من صفحة إنشاء الحساب")

    role_enum = RoleEnum.ADMIN if requested_role in ("admin", "principal", "super_admin") else RoleEnum.TEACHER
    
    new_user = User(
        school_id=school.id,
        public_id=generate_unique_public_id(db),
        email=request.email.lower(),
        full_name=request.full_name,
        specialization=(" ".join(str(request.specialization).strip().split()) if request.specialization else None),
        profile_image=(str(request.profile_image).strip() if request.profile_image else None),
        password_hash=hash_password(request.password),
        role=role_enum,
        role_label=requested_role,
        is_active=True,
        is_super_admin=is_super_admin_role
    )
    
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    return UserResponse(
        id=new_user.id,
        public_id=new_user.public_id,
        email=new_user.email,
        full_name=new_user.full_name,
        phone=new_user.phone,
        specialization=new_user.specialization,
        profile_image=new_user.profile_image,
        role=new_user.role.value,
        role_label=new_user.role_label,
        school_name=new_user.school.name if new_user.school else None,
        is_active=new_user.is_active,
        is_super_admin=new_user.is_super_admin,
        created_at=new_user.created_at
    )


@app.post("/auth/oauth/facebook", response_model=LoginResponse)
async def facebook_login(request: FacebookLoginRequest, db: Session = Depends(get_db)):
    """Simulated Facebook login: accept email + optional full_name and optional school_code.
    If the user exists, return token; otherwise create a new user under the provided
    school_code or the first active school.
    """
    email = request.email.lower()

    # Find existing user
    user = db.query(User).filter(User.email == email).first()
    if user:
        if not user.is_active:
            raise HTTPException(status_code=401, detail="الحساب غير فعال")
        token = create_access_token(user.id)
        return LoginResponse(
            token=token,
            user=UserResponse(
                id=user.id,
                public_id=user.public_id,
                email=user.email,
                full_name=user.full_name,
                phone=user.phone,
                specialization=user.specialization,
                profile_image=user.profile_image,
                role=user.role.value,
                role_label=user.role_label,
                school_name=user.school.name if user.school else None,
                is_active=user.is_active,
                is_super_admin=user.is_super_admin,
                created_at=user.created_at
            ),
            message="تم تسجيل الدخول بنجاح عبر فيسبوك"
        )

    # Create new user
    school = None
    if request.school_code:
        school = db.query(School).filter(School.code == request.school_code).first()

    if not school:
        # pick first available school
        school = db.query(School).first()

    if not school:
        raise HTTPException(status_code=400, detail="لا توجد مدرسة متاحة لإنشاء المستخدم")

    new_user = User(
        school_id=school.id,
        public_id=generate_unique_public_id(db),
        email=email,
        password_hash=hash_password('fb-login'),
        full_name=(request.full_name or email.split('@')[0]),
        role=RoleEnum.TEACHER,
        role_label="teacher",
        is_active=True,
        is_super_admin=False
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    token = create_access_token(new_user.id)
    return LoginResponse(
        token=token,
        user=UserResponse(
            id=new_user.id,
            public_id=new_user.public_id,
            email=new_user.email,
            full_name=new_user.full_name,
            phone=new_user.phone,
            specialization=new_user.specialization,
            profile_image=new_user.profile_image,
            role=new_user.role.value,
            role_label=new_user.role_label,
            school_name=new_user.school.name if new_user.school else None,
            is_active=new_user.is_active,
            is_super_admin=new_user.is_super_admin,
            created_at=new_user.created_at
        ),
        message="تم إنشاء الحساب وتسجيل الدخول عبر فيسبوك"
    )

# ===== Schools Routes =====

@app.get("/schools", response_model=SchoolsListResponse)
async def get_schools(db: Session = Depends(get_db)):
    """الحصول على قائمة جميع المدارس"""
    schools = db.query(School).order_by(School.name.asc()).all()
    return SchoolsListResponse(schools=schools, total=len(schools))

@app.get("/schools/{school_id}", response_model=SchoolResponse)
async def get_school(school_id: int, db: Session = Depends(get_db)):
    """الحصول على بيانات مدرسة معينة"""
    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="المدرسة غير موجودة")
    return school


@app.post("/schools", response_model=SchoolResponse)
async def create_school(
    request: SchoolCreateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """إنشاء مدرسة جديدة (مسؤول النظام فقط)"""
    ensure_super_admin_user(current_user)

    school_name = (request.name or "").strip()
    school_code = (request.code or "").strip().upper()

    if not school_name:
        raise HTTPException(status_code=400, detail="اسم المدرسة مطلوب")
    if not school_code:
        raise HTTPException(status_code=400, detail="كود المدرسة مطلوب")

    duplicate_name = db.query(School).filter(School.name == school_name).first()
    if duplicate_name:
        raise HTTPException(status_code=400, detail="اسم المدرسة مستخدم مسبقًا")

    duplicate_code = db.query(School).filter(School.code == school_code).first()
    if duplicate_code:
        raise HTTPException(status_code=400, detail="كود المدرسة مستخدم مسبقًا")

    school = School(name=school_name, code=school_code, status="active")
    db.add(school)
    db.commit()
    db.refresh(school)
    return school


@app.put("/schools/{school_id}", response_model=SchoolResponse)
async def update_school(
    school_id: int,
    request: SchoolUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """تعديل مدرسة (مسؤول النظام فقط)"""
    ensure_super_admin_user(current_user)

    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="المدرسة غير موجودة")

    school_name = (request.name or "").strip()
    school_code = (request.code or "").strip().upper()

    if not school_name:
        raise HTTPException(status_code=400, detail="اسم المدرسة مطلوب")
    if not school_code:
        raise HTTPException(status_code=400, detail="كود المدرسة مطلوب")

    duplicate_name = db.query(School).filter(
        School.name == school_name,
        School.id != school_id
    ).first()
    if duplicate_name:
        raise HTTPException(status_code=400, detail="اسم المدرسة مستخدم مسبقًا")

    duplicate_code = db.query(School).filter(
        School.code == school_code,
        School.id != school_id
    ).first()
    if duplicate_code:
        raise HTTPException(status_code=400, detail="كود المدرسة مستخدم مسبقًا")

    school.name = school_name
    school.code = school_code
    db.commit()
    db.refresh(school)
    return school


@app.delete("/schools/{school_id}")
async def delete_school(
    school_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """حذف مدرسة (مسؤول النظام فقط)"""
    ensure_super_admin_user(current_user)

    school = db.query(School).filter(School.id == school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="المدرسة غير موجودة")

    linked_users_count = db.query(func.count(User.id)).filter(User.school_id == school.id).scalar() or 0
    if linked_users_count > 0:
        raise HTTPException(
            status_code=409,
            detail=f"لا يمكن حذف المدرسة لأنها مرتبطة بـ {linked_users_count} مستخدم. احذف/انقل المستخدمين أولاً."
        )

    db.delete(school)
    db.commit()
    return {"message": "تم حذف المدرسة بنجاح"}


class SchoolApplicationRequest(BaseModel):
    name: str
    code: Optional[str] = None
    contact_email: Optional[EmailStr] = None
    contact_phone: Optional[str] = None
    notes: Optional[str] = None


@app.post("/schools/apply")
async def apply_school(request: SchoolApplicationRequest, db: Session = Depends(get_db)):
    """Submit a school application/request to join the system."""
    # Create application record
    from models import SchoolApplication

    app_rec = SchoolApplication(
        name=request.name,
        code=request.code,
        contact_email=request.contact_email,
        contact_phone=request.contact_phone,
        notes=request.notes,
        status="pending"
    )
    db.add(app_rec)
    db.commit()
    db.refresh(app_rec)

    return {"message": "تم تقديم طلب المدرسة بنجاح", "application_id": app_rec.id}


@app.get("/admin/school_applications")
async def list_school_applications(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """List pending school applications (super-admin only)."""
    if not current_user.is_super_admin:
        raise HTTPException(status_code=403, detail="غير مصرح")

    from models import SchoolApplication
    apps = db.query(SchoolApplication).order_by(SchoolApplication.created_at.desc()).all()
    return [
        {
            "id": a.id,
            "name": a.name,
            "code": a.code,
            "contact_email": a.contact_email,
            "contact_phone": a.contact_phone,
            "notes": a.notes,
            "status": a.status,
            "created_at": a.created_at
        }
        for a in apps
    ]


@app.post("/admin/school_applications/{app_id}/approve")
async def approve_school_application(app_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Approve a school application and create a School record. Super-admin only."""
    ensure_super_admin_user(current_user)

    from models import SchoolApplication
    app_rec = db.query(SchoolApplication).filter(SchoolApplication.id == app_id).first()
    if not app_rec:
        raise HTTPException(status_code=404, detail="طلب المدرسة غير موجود")

    # Create the school as active immediately
    existing = db.query(School).filter(School.name == app_rec.name).first()
    if existing:
        app_rec.status = "rejected"
        db.commit()
        raise HTTPException(status_code=400, detail="مدرسة بنفس الاسم موجودة بالفعل")

    candidate_code = (app_rec.code or f"SCH-{app_rec.id}").strip().upper()
    existing_code = db.query(School).filter(School.code == candidate_code).first()
    if existing_code:
        app_rec.status = "rejected"
        db.commit()
        raise HTTPException(status_code=400, detail="كود المدرسة مستخدم مسبقًا")

    new_school = School(name=app_rec.name, code=candidate_code, status="active")
    db.add(new_school)
    app_rec.status = "approved"
    db.commit()
    db.refresh(new_school)

    return {"message": "تمت الموافقة على المدرسة وأنشئت كسجل جديد", "school_id": new_school.id}

# ===== Classes Routes =====

@app.post("/admin/classes", response_model=ClassResponse)
async def admin_create_class(
    request: AdminClassUpsertRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """إنشاء صف وطلابه (إدارة فقط)"""
    ensure_admin_user(current_user)

    class_name = request.name.strip()
    if not class_name:
        raise HTTPException(status_code=400, detail="اسم الصف مطلوب")

    target_school_id = resolve_target_school_id(request.school_id, current_user)

    duplicate = db.query(Class).filter(
        Class.school_id == target_school_id,
        Class.name == class_name
    ).first()
    if duplicate:
        raise HTTPException(status_code=400, detail="الصف موجود بالفعل")

    new_class = Class(school_id=target_school_id, name=class_name)
    db.add(new_class)
    db.flush()

    for student_name in normalize_student_names(request.students):
        db.add(Student(class_id=new_class.id, full_name=student_name, is_active=True))

    db.commit()
    db.refresh(new_class)

    return ClassResponse(
        id=new_class.id,
        name=new_class.name,
        school_id=new_class.school_id,
        total_students=len([s for s in new_class.students if s.is_active])
    )


@app.put("/admin/classes/{class_id}", response_model=ClassResponse)
async def admin_update_class(
    class_id: int,
    request: AdminClassUpsertRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """تعديل اسم الصف وطلابه (إدارة فقط)"""
    ensure_admin_user(current_user)

    class_obj = db.query(Class).filter(Class.id == class_id).first()
    if not class_obj:
        raise HTTPException(status_code=404, detail="الصف غير موجود")

    target_school_id = resolve_target_school_id(request.school_id, current_user)
    if class_obj.school_id != target_school_id:
        raise HTTPException(status_code=403, detail="غير مصرح بتعديل هذا الصف")

    class_name = request.name.strip()
    if not class_name:
        raise HTTPException(status_code=400, detail="اسم الصف مطلوب")

    duplicate = db.query(Class).filter(
        Class.school_id == class_obj.school_id,
        Class.name == class_name,
        Class.id != class_obj.id
    ).first()
    if duplicate:
        raise HTTPException(status_code=400, detail="اسم الصف مستخدم مسبقًا")

    class_obj.name = class_name
    sync_students_for_class(db, class_obj, request.students)

    db.commit()
    db.refresh(class_obj)

    active_students = [s for s in class_obj.students if s.is_active]
    return ClassResponse(
        id=class_obj.id,
        name=class_obj.name,
        school_id=class_obj.school_id,
        total_students=len(active_students)
    )


@app.delete("/admin/classes/{class_id}")
async def admin_delete_class(
    class_id: int,
    school_id: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """حذف صف كامل مع طلابه (إدارة فقط)"""
    ensure_admin_user(current_user)

    class_obj = db.query(Class).filter(Class.id == class_id).first()
    if not class_obj:
        raise HTTPException(status_code=404, detail="الصف غير موجود")

    target_school_id = resolve_target_school_id(school_id, current_user)
    if class_obj.school_id != target_school_id:
        raise HTTPException(status_code=403, detail="غير مصرح بحذف هذا الصف")

    db.delete(class_obj)
    db.commit()
    return {"message": "تم حذف الصف بنجاح"}

@app.get("/classes", response_model=List[ClassResponse])
async def get_classes(
    school_id: Optional[int] = Query(None),
    school_code: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """الحصول على قائمة الصفوف. المشرف العام يستطيع تمرير `school_id` أو `school_code`."""
    target_school_id = None

    # Only super admin can request classes for any school by id or code
    if current_user.is_super_admin:
        if school_id:
            target_school_id = school_id
        elif school_code:
            sch = db.query(School).filter(School.code == school_code).first()
            if sch:
                target_school_id = sch.id
    else:
        target_school_id = current_user.school_id

    if not target_school_id:
        raise HTTPException(status_code=400, detail="رمز المدرسة أو معرفها مطلوب")

    classes = db.query(Class).filter(Class.school_id == target_school_id).order_by(Class.id).all()
    
    result = []
    for cls in classes:
        result.append(ClassResponse(
            id=cls.id,
            name=cls.name,
            school_id=cls.school_id,
            total_students=len(cls.students)
        ))
    
    return result

@app.get("/classes/{class_id}", response_model=ClassResponse)
async def get_class(
    class_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """الحصول على بيانات صف معين"""
    cls = db.query(Class).filter(Class.id == class_id).first()

    if not cls:
        raise HTTPException(status_code=404, detail="الصف غير موجود")

    # Non super admin users can only access classes in their own school
    if not current_user.is_super_admin and cls.school_id != current_user.school_id:
        raise HTTPException(status_code=403, detail="غير مصرح بالوصول لهذا الصف")
    
    return ClassResponse(
        id=cls.id,
        name=cls.name,
        school_id=cls.school_id,
        total_students=len(cls.students)
    )

# ===== Students Routes =====

@app.get("/classes/{class_id}/students")
async def get_class_students(
    class_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    search: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """الحصول على قائمة طلاب صف معين"""
    cls = db.query(Class).filter(Class.id == class_id).first()

    if not cls:
        raise HTTPException(status_code=404, detail="الصف غير موجود")

    if not current_user.is_super_admin and cls.school_id != current_user.school_id:
        raise HTTPException(status_code=403, detail="غير مصرح بالوصول لهذا الصف")
    
    students = db.query(Student).filter(
        and_(
            Student.class_id == class_id,
            Student.is_active == True
        )
    ).order_by(Student.id).all()
    
    return ClassStudentsResponse(
        class_id=cls.id,
        class_name=cls.name,
        total_students=len(students),
        students=[StudentResponse(id=s.id, full_name=s.full_name) for s in students]
    )


# Temporary debug endpoint to list all students (ignoring is_active)
@app.get("/debug/classes/{class_id}/students_all")
async def debug_get_all_students(
    class_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    cls = db.query(Class).filter(Class.id == class_id).first()
    if not cls:
        raise HTTPException(status_code=404, detail="الصف غير موجود")

    if not current_user.is_super_admin and cls.school_id != current_user.school_id:
        raise HTTPException(status_code=403, detail="غير مصرح بالوصول لهذا الصف")

    students = db.query(Student).filter(Student.class_id == class_id).order_by(Student.full_name).all()
    return {
        "class_id": cls.id,
        "class_name": cls.name,
        "total_students_all": len(students),
        "students_all": [{"id": s.id, "full_name": s.full_name, "is_active": getattr(s, 'is_active', None)} for s in students]
    }

# ===== Social/Search/Chat Routes =====

@app.get("/users/search", response_model=List[UserSearchItemResponse])
async def search_users(
    q: str = Query(..., min_length=1),
    school_id: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    target_school_id = resolve_target_school_id(school_id, current_user)
    query_text = q.strip()
    if not query_text:
        return []

    users_query = db.query(User).filter(
        User.id != current_user.id,
        User.is_active == True,
        User.school_id == target_school_id,
    )

    if query_text.isdigit() and len(query_text) == 5:
        users_query = users_query.filter(User.public_id == query_text)
    else:
        users_query = users_query.filter(User.full_name.ilike(f"%{query_text}%"))

    users = users_query.order_by(User.full_name.asc()).all()

    return [
        UserSearchItemResponse(
            id=u.id,
            public_id=u.public_id,
            full_name=u.full_name,
            email=u.email,
            specialization=u.specialization,
            profile_image=u.profile_image,
            role=u.role.value,
            role_label=u.role_label or u.role.value,
        )
        for u in users if not is_blocked_between(db, current_user.id, u.id)
    ]


@app.get("/users/{user_id}/profile", response_model=UserProfileCardResponse)
async def get_user_profile_card(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.id == user_id, User.is_active == True).first()
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")

    if not current_user.is_super_admin and user.school_id != current_user.school_id and user.id != current_user.id:
        raise HTTPException(status_code=403, detail="غير مصرح")

    return UserProfileCardResponse(
        id=user.id,
        public_id=user.public_id,
        full_name=user.full_name,
        email=user.email,
        phone=user.phone,
        specialization=user.specialization,
        profile_image=user.profile_image,
        school_name=user.school.name if user.school else None,
        role_label=user.role_label or user.role.value,
    )


@app.post("/users/{user_id}/block")
async def block_user(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="لا يمكن حظر نفسك")

    target = db.query(User).filter(User.id == user_id, User.is_active == True).first()
    if not target:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")

    existing = db.query(UserBlock).filter(
        UserBlock.blocker_id == current_user.id,
        UserBlock.blocked_id == user_id,
    ).first()
    if existing:
        return {"message": "المستخدم محظور مسبقًا"}

    block = UserBlock(blocker_id=current_user.id, blocked_id=user_id)
    db.add(block)

    fr = get_friendship_between(db, current_user.id, user_id)
    if fr:
        db.delete(fr)

    sessions = db.query(ChatSession).filter(
        or_(
            and_(ChatSession.starter_id == current_user.id, ChatSession.joiner_id == user_id),
            and_(ChatSession.starter_id == user_id, ChatSession.joiner_id == current_user.id),
        )
    ).all()
    for session_obj in sessions:
        db.query(Message).filter(Message.session_id == session_obj.id).delete()
        db.query(ChatMute).filter(ChatMute.session_id == session_obj.id).delete()
        db.delete(session_obj)

    db.commit()
    return {"message": "تم حظر المستخدم بنجاح"}


@app.delete("/users/{user_id}/block")
async def unblock_user(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    block = db.query(UserBlock).filter(
        UserBlock.blocker_id == current_user.id,
        UserBlock.blocked_id == user_id,
    ).first()
    if not block:
        return {"message": "المستخدم غير محظور"}

    db.delete(block)
    db.commit()
    return {"message": "تم إلغاء حظر المستخدم"}


@app.get("/users/blocked", response_model=List[BlockedUserItemResponse])
async def list_blocked_users(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    blocks = db.query(UserBlock).filter(
        UserBlock.blocker_id == current_user.id,
    ).order_by(UserBlock.created_at.desc()).all()

    result = []
    for block in blocks:
        blocked_user = db.query(User).filter(User.id == block.blocked_id, User.is_active == True).first()
        if not blocked_user:
            continue
        result.append(BlockedUserItemResponse(
            user_id=blocked_user.id,
            public_id=blocked_user.public_id,
            full_name=blocked_user.full_name,
            specialization=blocked_user.specialization,
            profile_image=blocked_user.profile_image,
            role_label=blocked_user.role_label or blocked_user.role.value,
            blocked_at=block.created_at,
        ))
    return result


@app.delete("/admin/users/{user_id}")
async def admin_delete_user(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user.is_super_admin:
        raise HTTPException(status_code=403, detail="غير مصرح")

    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="لا يمكن حذف حسابك الحالي")

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")

    db.query(Friendship).filter(
        or_(Friendship.sender_id == user_id, Friendship.receiver_id == user_id)
    ).delete()
    db.query(UserBlock).filter(
        or_(UserBlock.blocker_id == user_id, UserBlock.blocked_id == user_id)
    ).delete()

    user_sessions = db.query(ChatSession).filter(
        or_(ChatSession.starter_id == user_id, ChatSession.joiner_id == user_id)
    ).all()
    for session_obj in user_sessions:
        db.query(Message).filter(Message.session_id == session_obj.id).delete()
        db.query(ChatMute).filter(ChatMute.session_id == session_obj.id).delete()
        db.delete(session_obj)

    db.delete(user)
    db.commit()
    return {"message": "تم حذف الحساب بنجاح"}


@app.put("/admin/users/{user_id}/role")
async def admin_update_user_role(
    user_id: int,
    request: AdminUserRoleUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user.is_super_admin:
        raise HTTPException(status_code=403, detail="غير مصرح")

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")

    role_label = (request.role_label or "").strip().lower()
    if role_label not in ("teacher", "admin", "principal", "super_admin"):
        raise HTTPException(status_code=400, detail="نوع الحساب غير مدعوم")

    user.role_label = role_label
    user.role = RoleEnum.ADMIN if role_label in ("admin", "principal", "super_admin") else RoleEnum.TEACHER
    user.is_super_admin = role_label == "super_admin"
    if user.is_super_admin:
        user.email = user.email or LAUNCH_SUPER_ADMIN_EMAIL
    db.commit()
    db.refresh(user)

    return {
        "message": "تم تحديث صلاحية المستخدم",
        "user_id": user.id,
        "role": user.role.value,
        "role_label": user.role_label,
        "is_super_admin": user.is_super_admin,
    }


@app.get("/friends/list")
async def get_friends_list(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    friendships = db.query(Friendship).filter(
        Friendship.status == FriendshipStatusEnum.ACCEPTED,
        or_(Friendship.sender_id == current_user.id, Friendship.receiver_id == current_user.id),
    ).order_by(Friendship.updated_at.desc()).all()

    items = []
    for fr in friendships:
        other_id = fr.receiver_id if fr.sender_id == current_user.id else fr.sender_id
        if is_blocked_between(db, current_user.id, other_id):
            continue
        other_user = db.query(User).filter(User.id == other_id, User.is_active == True).first()
        if not other_user:
            continue
        items.append({
            "user_id": other_user.id,
            "public_id": other_user.public_id,
            "full_name": other_user.full_name,
            "specialization": other_user.specialization,
            "profile_image": other_user.profile_image,
            "role_label": (other_user.role_label or other_user.role.value),
            "friendship_id": fr.id,
            "status": fr.status.value,
        })

    return {"friends": items}


@app.get("/friends/requests/incoming", response_model=List[FriendRequestItemResponse])
async def get_incoming_friend_requests(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    requests = db.query(Friendship).filter(
        Friendship.receiver_id == current_user.id,
        Friendship.status == FriendshipStatusEnum.PENDING,
    ).order_by(Friendship.created_at.desc()).all()

    result = []
    for fr in requests:
        if is_blocked_between(db, current_user.id, fr.sender_id):
            continue
        sender = db.query(User).filter(User.id == fr.sender_id).first()
        if not sender:
            continue
        result.append(FriendRequestItemResponse(
            request_id=fr.id,
            user_id=sender.id,
            public_id=sender.public_id,
            full_name=sender.full_name,
            specialization=sender.specialization,
            profile_image=sender.profile_image,
            role_label=sender.role_label or sender.role.value,
            status=fr.status.value,
            created_at=fr.created_at,
        ))
    return result


@app.get("/friends/requests/outgoing", response_model=List[FriendRequestItemResponse])
async def get_outgoing_friend_requests(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    requests = db.query(Friendship).filter(
        Friendship.sender_id == current_user.id,
        Friendship.status == FriendshipStatusEnum.PENDING,
    ).order_by(Friendship.created_at.desc()).all()

    result = []
    for fr in requests:
        if is_blocked_between(db, current_user.id, fr.receiver_id):
            continue
        receiver = db.query(User).filter(User.id == fr.receiver_id).first()
        if not receiver:
            continue
        result.append(FriendRequestItemResponse(
            request_id=fr.id,
            user_id=receiver.id,
            public_id=receiver.public_id,
            full_name=receiver.full_name,
            specialization=receiver.specialization,
            profile_image=receiver.profile_image,
            role_label=receiver.role_label or receiver.role.value,
            status=fr.status.value,
            created_at=fr.created_at,
        ))
    return result


@app.post("/friends/requests")
async def create_friend_request(
    request: FriendRequestCreateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if request.receiver_id == current_user.id:
        raise HTTPException(status_code=400, detail="لا يمكن إرسال طلب إضافة لنفسك")

    receiver = db.query(User).filter(User.id == request.receiver_id, User.is_active == True).first()
    if not receiver:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")

    if is_blocked_between(db, current_user.id, receiver.id):
        raise HTTPException(status_code=403, detail="لا يمكن إرسال طلب إضافة بسبب الحظر")

    if not current_user.is_super_admin and receiver.school_id != current_user.school_id:
        raise HTTPException(status_code=403, detail="يمكن الإضافة داخل المدرسة نفسها فقط")

    existing = get_friendship_between(db, current_user.id, receiver.id)
    if existing:
        if existing.status == FriendshipStatusEnum.ACCEPTED:
            return {"message": "المستخدم مضاف بالفعل", "request_id": existing.id}
        if existing.status == FriendshipStatusEnum.PENDING:
            if existing.sender_id == current_user.id:
                return {"message": "تم إرسال الطلب مسبقًا", "request_id": existing.id}
            existing.status = FriendshipStatusEnum.ACCEPTED
            existing.updated_at = datetime.utcnow()
            db.commit()
            return {"message": "تم قبول طلب الإضافة المتبادل تلقائيًا", "request_id": existing.id}

        existing.status = FriendshipStatusEnum.PENDING
        existing.sender_id = current_user.id
        existing.receiver_id = receiver.id
        existing.updated_at = datetime.utcnow()
        db.commit()
        return {"message": "تم إرسال طلب الإضافة", "request_id": existing.id}

    fr = Friendship(
        sender_id=current_user.id,
        receiver_id=receiver.id,
        status=FriendshipStatusEnum.PENDING,
    )
    db.add(fr)
    db.commit()
    db.refresh(fr)
    return {"message": "تم إرسال طلب الإضافة", "request_id": fr.id}


@app.post("/friends/requests/{request_id}/respond")
async def respond_friend_request(
    request_id: int,
    request: FriendRequestRespondRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    fr = db.query(Friendship).filter(Friendship.id == request_id).first()
    if not fr:
        raise HTTPException(status_code=404, detail="طلب الإضافة غير موجود")

    if fr.receiver_id != current_user.id:
        raise HTTPException(status_code=403, detail="غير مصرح")

    if fr.status != FriendshipStatusEnum.PENDING:
        raise HTTPException(status_code=400, detail="تم التعامل مع الطلب مسبقًا")

    action = (request.action or "").strip().lower()
    if action not in ("accept", "reject"):
        raise HTTPException(status_code=400, detail="الإجراء يجب أن يكون accept أو reject")

    fr.status = FriendshipStatusEnum.ACCEPTED if action == "accept" else FriendshipStatusEnum.REJECTED
    fr.updated_at = datetime.utcnow()
    db.commit()

    return {
        "message": "تم قبول الطلب" if action == "accept" else "تم رفض الطلب",
        "request_id": fr.id,
        "status": fr.status.value,
    }


@app.delete("/friends/requests/{request_id}")
async def cancel_friend_request(
    request_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    fr = db.query(Friendship).filter(Friendship.id == request_id).first()
    if not fr:
        raise HTTPException(status_code=404, detail="طلب الإضافة غير موجود")

    if fr.sender_id != current_user.id:
        raise HTTPException(status_code=403, detail="لا يمكنك إلغاء هذا الطلب")

    if fr.status != FriendshipStatusEnum.PENDING:
        raise HTTPException(status_code=400, detail="لا يمكن إلغاء الطلب بعد معالجته")

    db.delete(fr)
    db.commit()
    return {
        "message": "تم إلغاء طلب الصداقة",
        "request_id": request_id,
    }


@app.delete("/friends/{friend_user_id}")
async def remove_friend(
    friend_user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if friend_user_id == current_user.id:
        raise HTTPException(status_code=400, detail="لا يمكن إزالة نفسك من الأصدقاء")

    friend = db.query(User).filter(User.id == friend_user_id, User.is_active == True).first()
    if not friend:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")

    friendship = get_friendship_between(db, current_user.id, friend_user_id)
    if not friendship or friendship.status != FriendshipStatusEnum.ACCEPTED:
        raise HTTPException(status_code=404, detail="المستخدم ليس ضمن قائمة أصدقائك")

    friendship.status = FriendshipStatusEnum.REJECTED
    friendship.updated_at = datetime.utcnow()
    db.commit()

    return {
        "message": "تمت إزالة المستخدم من قائمة الأصدقاء",
        "friend_user_id": friend_user_id,
    }


@app.get("/chat/sessions")
async def list_chat_sessions(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    sessions = db.query(ChatSession).filter(
        or_(ChatSession.starter_id == current_user.id, ChatSession.joiner_id == current_user.id)
    ).order_by(ChatSession.updated_at.desc()).all()

    result = []
    for session_obj in sessions:
        partner_id = session_obj.joiner_id if session_obj.starter_id == current_user.id else session_obj.starter_id
        partner = db.query(User).filter(User.id == partner_id, User.is_active == True).first()
        if not partner:
            continue

        if is_blocked_between(db, current_user.id, partner.id):
            continue

        fr = get_friendship_between(db, current_user.id, partner.id)
        if not fr or fr.status != FriendshipStatusEnum.ACCEPTED:
            continue

        last_message = db.query(Message).filter(Message.session_id == session_obj.id).order_by(Message.id.desc()).first()
        mute_setting = db.query(ChatMute).filter(
            ChatMute.user_id == current_user.id,
            ChatMute.session_id == session_obj.id,
        ).first()
        result.append({
            "session_id": session_obj.id,
            "partner_id": partner.id,
            "partner_public_id": partner.public_id,
            "partner_name": partner.full_name,
            "partner_profile_image": partner.profile_image,
            "partner_role_label": partner.role_label or partner.role.value,
            "partner_specialization": partner.specialization,
            "last_message": last_message.content if last_message else "",
            "last_message_at": last_message.sent_at.isoformat() if last_message else None,
            "last_message_sender_id": last_message.sender_id if last_message else None,
            "is_muted": bool(mute_setting and mute_setting.is_muted),
        })

    return {"sessions": result}


@app.post("/chat/sessions/with/{user_id}")
async def get_or_create_chat_with_user(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="لا يمكن فتح دردشة مع نفسك")

    partner = db.query(User).filter(User.id == user_id, User.is_active == True).first()
    if not partner:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")

    if is_blocked_between(db, current_user.id, user_id):
        raise HTTPException(status_code=403, detail="لا يمكن بدء الدردشة بسبب الحظر")

    fr = get_friendship_between(db, current_user.id, user_id)
    if not fr or fr.status != FriendshipStatusEnum.ACCEPTED:
        raise HTTPException(status_code=403, detail="لا يمكن بدء الدردشة قبل قبول طلب الإضافة")

    session_obj = get_or_create_chat_session(db, current_user.id, user_id)
    db.commit()
    db.refresh(session_obj)

    return {
        "session_id": session_obj.id,
        "partner_id": partner.id,
        "partner_public_id": partner.public_id,
        "partner_name": partner.full_name,
        "partner_role_label": partner.role_label or partner.role.value,
    }


@app.get("/chat/sessions/{session_id}/messages")
async def get_chat_messages(
    session_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    session_obj = db.query(ChatSession).filter(ChatSession.id == session_id).first()
    if not session_obj:
        raise HTTPException(status_code=404, detail="جلسة الدردشة غير موجودة")

    if current_user.id not in (session_obj.starter_id, session_obj.joiner_id):
        raise HTTPException(status_code=403, detail="غير مصرح")

    partner_id = session_obj.joiner_id if session_obj.starter_id == current_user.id else session_obj.starter_id
    if is_blocked_between(db, current_user.id, partner_id):
        raise HTTPException(status_code=403, detail="لا يمكن الوصول للدردشة بسبب الحظر")
    fr = get_friendship_between(db, current_user.id, partner_id)
    if not fr or fr.status != FriendshipStatusEnum.ACCEPTED:
        raise HTTPException(status_code=403, detail="الدردشة متاحة بعد قبول طلب الإضافة فقط")

    messages = db.query(Message).filter(Message.session_id == session_id).order_by(Message.id.asc()).all()
    return {
        "messages": [
            {
                "id": m.id,
                "sender_id": m.sender_id,
                "content": m.content,
                "sent_at": m.sent_at.isoformat(),
            }
            for m in messages
        ]
    }


@app.post("/chat/sessions/{session_id}/messages")
async def send_chat_message(
    session_id: int,
    request: ChatMessageCreateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    session_obj = db.query(ChatSession).filter(ChatSession.id == session_id).first()
    if not session_obj:
        raise HTTPException(status_code=404, detail="جلسة الدردشة غير موجودة")

    if current_user.id not in (session_obj.starter_id, session_obj.joiner_id):
        raise HTTPException(status_code=403, detail="غير مصرح")

    content = (request.content or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="نص الرسالة مطلوب")

    partner_id = session_obj.joiner_id if session_obj.starter_id == current_user.id else session_obj.starter_id
    if is_blocked_between(db, current_user.id, partner_id):
        raise HTTPException(status_code=403, detail="لا يمكن إرسال رسالة بسبب الحظر")
    fr = get_friendship_between(db, current_user.id, partner_id)
    if not fr or fr.status != FriendshipStatusEnum.ACCEPTED:
        raise HTTPException(status_code=403, detail="الدردشة متاحة بعد قبول طلب الإضافة فقط")

    message = Message(session_id=session_obj.id, sender_id=current_user.id, content=content)
    session_obj.updated_at = datetime.utcnow()
    db.add(message)
    db.commit()
    db.refresh(message)

    return {
        "id": message.id,
        "sender_id": message.sender_id,
        "content": message.content,
        "sent_at": message.sent_at.isoformat(),
    }


@app.post("/chat/sessions/{session_id}/mute")
async def set_chat_mute(
    session_id: int,
    request: ChatMuteToggleRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    session_obj = db.query(ChatSession).filter(ChatSession.id == session_id).first()
    if not session_obj:
        raise HTTPException(status_code=404, detail="جلسة الدردشة غير موجودة")

    if current_user.id not in (session_obj.starter_id, session_obj.joiner_id):
        raise HTTPException(status_code=403, detail="غير مصرح")

    mute_record = db.query(ChatMute).filter(
        ChatMute.user_id == current_user.id,
        ChatMute.session_id == session_id,
    ).first()

    if not mute_record:
        mute_record = ChatMute(user_id=current_user.id, session_id=session_id, is_muted=bool(request.is_muted))
        db.add(mute_record)
    else:
        mute_record.is_muted = bool(request.is_muted)

    db.commit()
    return {"message": "تم تحديث حالة كتم الإشعارات", "is_muted": bool(request.is_muted)}


@app.delete("/chat/sessions/{session_id}")
async def delete_chat_session(
    session_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    session_obj = db.query(ChatSession).filter(ChatSession.id == session_id).first()
    if not session_obj:
        raise HTTPException(status_code=404, detail="جلسة الدردشة غير موجودة")

    if current_user.id not in (session_obj.starter_id, session_obj.joiner_id):
        raise HTTPException(status_code=403, detail="غير مصرح")

    db.query(Message).filter(Message.session_id == session_id).delete()
    session_obj.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "تم حذف جميع رسائل الدردشة"}


@app.put("/chat/messages/{message_id}")
async def edit_chat_message(
    message_id: int,
    request: ChatMessageEditRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    message = db.query(Message).filter(Message.id == message_id).first()
    if not message:
        raise HTTPException(status_code=404, detail="الرسالة غير موجودة")

    if message.sender_id != current_user.id:
        raise HTTPException(status_code=403, detail="لا يمكنك تعديل رسالة شخص آخر")

    content = (request.content or "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="نص الرسالة مطلوب")

    message.content = content
    db.commit()
    db.refresh(message)
    return {
        "id": message.id,
        "sender_id": message.sender_id,
        "content": message.content,
        "sent_at": message.sent_at.isoformat(),
        "edited": True,
    }


@app.delete("/chat/messages/{message_id}")
async def delete_chat_message(
    message_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    message = db.query(Message).filter(Message.id == message_id).first()
    if not message:
        raise HTTPException(status_code=404, detail="الرسالة غير موجودة")

    if message.sender_id != current_user.id:
        raise HTTPException(status_code=403, detail="لا يمكنك حذف رسالة شخص آخر")

    session_obj = db.query(ChatSession).filter(ChatSession.id == message.session_id).first()
    db.delete(message)
    if session_obj:
        session_obj.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "تم حذف الرسالة", "id": message_id}


@app.post("/users/status/online")
async def heartbeat_online(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    current_user.last_seen = datetime.utcnow()
    db.commit()
    return {"status": "ok"}


@app.get("/users/{user_id}/status")
async def get_user_status(
    user_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.id == user_id, User.is_active == True).first()
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    if user.last_seen is None:
        return {"user_id": user_id, "online": False, "last_seen": None}
    delta = datetime.utcnow() - user.last_seen
    return {
        "user_id": user_id,
        "online": delta.total_seconds() < 120,
        "last_seen": user.last_seen.isoformat(),
    }


@app.post("/calls/start")
async def start_call(
    request: CallStartRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    callee = db.query(User).filter(User.id == request.callee_id, User.is_active == True).first()
    if not callee:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    if callee.id == current_user.id:
        raise HTTPException(status_code=400, detail="لا يمكن الاتصال بنفسك")
    if is_blocked_between(db, current_user.id, callee.id):
        raise HTTPException(status_code=403, detail="لا يمكن بدء الاتصال بسبب الحظر")

    friendship = get_friendship_between(db, current_user.id, callee.id)
    if not friendship or friendship.status != FriendshipStatusEnum.ACCEPTED:
        raise HTTPException(status_code=403, detail="الاتصال متاح بعد قبول الصداقة فقط")

    active_call = db.query(CallSession).filter(
        or_(
            and_(CallSession.caller_id == current_user.id, CallSession.callee_id == callee.id),
            and_(CallSession.caller_id == callee.id, CallSession.callee_id == current_user.id),
        ),
        CallSession.status.in_(["ringing", "answered"])
    ).order_by(CallSession.id.desc()).first()
    if active_call:
        return {"call_id": active_call.id, "status": active_call.status}

    call = CallSession(caller_id=current_user.id, callee_id=callee.id, status="ringing")
    db.add(call)
    db.commit()
    db.refresh(call)

    return {"call_id": call.id, "status": call.status}


@app.get("/calls/incoming")
async def get_incoming_call(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    call = db.query(CallSession).filter(
        CallSession.callee_id == current_user.id,
        CallSession.status == "ringing"
    ).order_by(CallSession.id.desc()).first()
    if not call:
        return {"incoming": None}

    caller = db.query(User).filter(User.id == call.caller_id).first()
    return {
        "incoming": {
            "call_id": call.id,
            "caller_id": call.caller_id,
            "caller_name": caller.full_name if caller else "-",
            "created_at": call.created_at.isoformat() if call.created_at else None,
        }
    }


@app.get("/calls/{call_id}")
async def get_call_state(
    call_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    call = db.query(CallSession).filter(CallSession.id == call_id).first()
    if not call:
        raise HTTPException(status_code=404, detail="المكالمة غير موجودة")
    if current_user.id not in (call.caller_id, call.callee_id):
        raise HTTPException(status_code=403, detail="غير مصرح")

    return {
        "call_id": call.id,
        "caller_id": call.caller_id,
        "callee_id": call.callee_id,
        "status": call.status,
        "offer_sdp": call.offer_sdp,
        "answer_sdp": call.answer_sdp,
    }


@app.post("/calls/{call_id}/offer")
async def set_call_offer(
    call_id: int,
    request: CallSdpRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    call = db.query(CallSession).filter(CallSession.id == call_id).first()
    if not call:
        raise HTTPException(status_code=404, detail="المكالمة غير موجودة")
    if current_user.id != call.caller_id:
        raise HTTPException(status_code=403, detail="فقط المتصل يمكنه إرسال العرض")
    if call.status not in ("ringing", "answered"):
        raise HTTPException(status_code=409, detail="المكالمة منتهية")

    call.offer_sdp = request.sdp
    call.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "تم حفظ عرض المكالمة"}


@app.post("/calls/{call_id}/answer")
async def set_call_answer(
    call_id: int,
    request: CallSdpRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    call = db.query(CallSession).filter(CallSession.id == call_id).first()
    if not call:
        raise HTTPException(status_code=404, detail="المكالمة غير موجودة")
    if current_user.id != call.callee_id:
        raise HTTPException(status_code=403, detail="فقط المستقبل يمكنه إرسال الإجابة")
    if call.status not in ("ringing", "answered"):
        raise HTTPException(status_code=409, detail="المكالمة منتهية")

    call.answer_sdp = request.sdp
    call.status = "answered"
    call.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "تم حفظ إجابة المكالمة"}


@app.post("/calls/{call_id}/ice")
async def push_call_ice_candidate(
    call_id: int,
    request: CallIceRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    call = db.query(CallSession).filter(CallSession.id == call_id).first()
    if not call:
        raise HTTPException(status_code=404, detail="المكالمة غير موجودة")
    if current_user.id not in (call.caller_id, call.callee_id):
        raise HTTPException(status_code=403, detail="غير مصرح")

    recipient_id = call.callee_id if current_user.id == call.caller_id else call.caller_id
    ice = CallIceCandidate(
        call_id=call.id,
        sender_id=current_user.id,
        recipient_id=recipient_id,
        candidate=request.candidate,
        sdp_mid=request.sdp_mid,
        sdp_mline_index=request.sdp_mline_index,
    )
    db.add(ice)
    db.commit()
    return {"message": "تم حفظ ICE candidate"}


@app.get("/calls/{call_id}/ice")
async def poll_call_ice_candidates(
    call_id: int,
    since_id: int = Query(0),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    call = db.query(CallSession).filter(CallSession.id == call_id).first()
    if not call:
        raise HTTPException(status_code=404, detail="المكالمة غير موجودة")
    if current_user.id not in (call.caller_id, call.callee_id):
        raise HTTPException(status_code=403, detail="غير مصرح")

    candidates = db.query(CallIceCandidate).filter(
        CallIceCandidate.call_id == call_id,
        CallIceCandidate.recipient_id == current_user.id,
        CallIceCandidate.id > since_id,
    ).order_by(CallIceCandidate.id.asc()).all()

    return {
        "candidates": [
            {
                "id": c.id,
                "candidate": c.candidate,
                "sdp_mid": c.sdp_mid,
                "sdp_mline_index": c.sdp_mline_index,
            }
            for c in candidates
        ]
    }


@app.post("/calls/{call_id}/end")
async def end_call(
    call_id: int,
    request: CallEndRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    call = db.query(CallSession).filter(CallSession.id == call_id).first()
    if not call:
        raise HTTPException(status_code=404, detail="المكالمة غير موجودة")
    if current_user.id not in (call.caller_id, call.callee_id):
        raise HTTPException(status_code=403, detail="غير مصرح")

    status = (request.status or "ended").strip().lower()
    if status not in ("ended", "rejected", "missed"):
        status = "ended"

    call.status = status
    call.ended_at = datetime.utcnow()
    call.updated_at = datetime.utcnow()
    db.commit()
    return {"message": "تم إنهاء المكالمة", "status": call.status}


# ===== Attendance Routes =====

@app.post("/attendance/submit", response_model=AttendanceSubmissionResponse)
async def submit_attendance(
    request: AttendanceSubmissionRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """تقديم تقرير حضور جديد"""
    purge_expired_archived_submissions(db)
    
    # Verify user is teacher and class belongs to their school
    cls = db.query(Class).filter(Class.id == request.class_id).first()

    if not cls:
        raise HTTPException(status_code=404, detail="الصف غير موجود")

    if not current_user.is_super_admin and cls.school_id != current_user.school_id:
        raise HTTPException(status_code=403, detail="غير مصرح بالوصول لهذا الصف")

    school_settings = ensure_school_term_settings(db, cls.school_id)
    
    existing_submission = db.query(AttendanceSubmission).filter(
        AttendanceSubmission.class_id == request.class_id,
        AttendanceSubmission.date == request.date,
        AttendanceSubmission.academic_year == school_settings.current_academic_year,
        AttendanceSubmission.term == school_settings.current_term,
        AttendanceSubmission.deleted_at.is_(None)
    ).first()
    if existing_submission:
        raise HTTPException(status_code=409, detail="تم تسجيل حضور هذا الصف لهذا اليوم مسبقاً.")

    # Create submission
    submission = AttendanceSubmission(
        class_id=request.class_id,
        date=request.date,
        submission_type=request.submission_type,
        num_sessions=request.num_sessions,
        academic_year=school_settings.current_academic_year,
        term=school_settings.current_term,
        submitted_by=current_user.id,
        status=SubmissionStatusEnum.SUBMITTED,
        submitted_at=datetime.utcnow()
    )
    
    db.add(submission)
    db.flush()  # Get submission ID
    
    # Add attendance records
    for record_req in request.records:
        # Verify student exists in class
        student = db.query(Student).filter(
            and_(
                Student.id == record_req.student_id,
                Student.class_id == request.class_id
            )
        ).first()
        
        if not student:
            db.rollback()
            raise HTTPException(status_code=404, detail=f"الطالب {record_req.student_id} غير موجود في الصف")
        
        record = AttendanceRecord(
            student_id=record_req.student_id,
            submission_id=submission.id,
            date=request.date,
            session_number=record_req.session_number,
            status=parse_binary_attendance_status(record_req.status),
            notes=record_req.notes,
            created_by=current_user.id
        )
        
        db.add(record)
    
    db.commit()
    db.refresh(submission)

    daily_classes_count = db.query(func.count(AttendanceSubmission.id)).join(
        Class, AttendanceSubmission.class_id == Class.id
    ).filter(
        Class.school_id == cls.school_id,
        AttendanceSubmission.date == request.date,
        AttendanceSubmission.academic_year == school_settings.current_academic_year,
        AttendanceSubmission.term == school_settings.current_term,
        AttendanceSubmission.deleted_at.is_(None)
    ).scalar() or 0
    
    return AttendanceSubmissionResponse(
        id=submission.id,
        class_id=submission.class_id,
        date=submission.date,
        daily_classes_count=int(daily_classes_count),
        submission_type=submission.submission_type,
        num_sessions=submission.num_sessions,
        academic_year=submission.academic_year,
        term=submission.term,
        status=submission.status.value,
        created_at=submission.created_at,
        submitted_at=submission.submitted_at,
        records=[
            AttendanceRecordResponse(
                id=r.id,
                student_id=r.student_id,
                date=r.date,
                session_number=r.session_number,
                status=r.status.value,
                notes=r.notes,
                created_at=r.created_at
            )
            for r in submission.records
        ]
    )

@app.get("/attendance/submissions", response_model=List[AttendanceSubmissionSummaryResponse])
async def get_submissions(
    class_id: Optional[int] = Query(None),
    school_id: Optional[int] = Query(None),
    academic_year: Optional[int] = Query(None),
    term: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    submission_date: Optional[date] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """الحصول على قائمة التقديمات"""
    purge_expired_archived_submissions(db)

    target_school_id = current_user.school_id
    if school_id and current_user.is_super_admin:
        target_school_id = school_id

    term_settings = ensure_school_term_settings(db, target_school_id)
    selected_year = academic_year if academic_year is not None else term_settings.current_academic_year
    selected_term = normalize_term_or_raise(term) if term else term_settings.current_term

    query = db.query(AttendanceSubmission).join(Class).filter(
        Class.school_id == target_school_id,
        AttendanceSubmission.deleted_at.is_(None),
        AttendanceSubmission.academic_year == selected_year,
        AttendanceSubmission.term == selected_term
    )
    
    if class_id:
        query = query.filter(AttendanceSubmission.class_id == class_id)
    
    if date_from:
        query = query.filter(AttendanceSubmission.date >= date_from)
    
    if date_to:
        query = query.filter(AttendanceSubmission.date <= date_to)

    if submission_date:
        query = query.filter(AttendanceSubmission.date == submission_date)
    
    submissions = query.order_by(AttendanceSubmission.date.desc()).all()
    
    result = []
    for sub in submissions:
        present_count = sum(1 for r in sub.records if r.status == AttendanceStatusEnum.PRESENT)
        absent_count = sum(1 for r in sub.records if r.status == AttendanceStatusEnum.ABSENT)

        result.append(AttendanceSubmissionSummaryResponse(
            id=sub.id,
            class_id=sub.class_id,
            teacher_name=sub.submitted_by_user.full_name if sub.submitted_by_user else "-",
            class_name=sub.class_obj.name if sub.class_obj else "-",
            date=sub.date,
            academic_year=sub.academic_year,
            term=sub.term,
            total_students=len(sub.records),
            present_count=present_count,
            absent_count=absent_count
        ))
    
    return result


@app.get("/attendance/submission-dates")
async def get_attendance_submission_dates(
    school_id: Optional[int] = Query(None),
    class_id: Optional[int] = Query(None),
    academic_year: Optional[int] = Query(None),
    term: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    purge_expired_archived_submissions(db)

    target_school_id = current_user.school_id
    if school_id and current_user.is_super_admin:
        target_school_id = school_id

    settings = ensure_school_term_settings(db, target_school_id)
    selected_year = academic_year if academic_year is not None else settings.current_academic_year
    selected_term = normalize_term_or_raise(term) if term else settings.current_term

    query = db.query(AttendanceSubmission.date).join(Class).filter(
        Class.school_id == target_school_id,
        AttendanceSubmission.deleted_at.is_(None),
        AttendanceSubmission.academic_year == selected_year,
        AttendanceSubmission.term == selected_term,
    )

    if class_id:
        query = query.filter(AttendanceSubmission.class_id == class_id)

    dates = query.order_by(AttendanceSubmission.date.desc()).all()
    unique_dates: List[str] = []
    seen = set()
    for row in dates:
        date_value = row[0].isoformat()
        if date_value in seen:
            continue
        seen.add(date_value)
        unique_dates.append(date_value)

    return {
        "school_id": target_school_id,
        "academic_year": selected_year,
        "term": selected_term,
        "dates": unique_dates,
    }


@app.get("/attendance/submissions/{submission_id}/details", response_model=AttendanceSubmissionDetailResponse)
async def get_submission_details(
    submission_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """تفاصيل سجل حضور معين مع أسماء الطلاب وحالتهم"""
    purge_expired_archived_submissions(db)
    submission = db.query(AttendanceSubmission).join(Class).filter(
        AttendanceSubmission.id == submission_id,
        AttendanceSubmission.deleted_at.is_(None)
    ).first()
    if not submission:
        raise HTTPException(status_code=404, detail="السجل غير موجود")

    if not current_user.is_super_admin and submission.class_obj.school_id != current_user.school_id:
        raise HTTPException(status_code=403, detail="غير مصرح")

    student_details = []
    for record in sorted(submission.records, key=lambda r: (r.student.full_name if r.student else "", r.id)):
        student_details.append(SubmissionStudentDetailResponse(
            record_id=record.id,
            student_id=record.student_id,
            student_name=record.student.full_name if record.student else "-",
            status=record.status.value
        ))

    return AttendanceSubmissionDetailResponse(
        submission_id=submission.id,
        teacher_name=submission.submitted_by_user.full_name if submission.submitted_by_user else "-",
        class_name=submission.class_obj.name if submission.class_obj else "-",
        date=submission.date,
        students=student_details
    )


@app.get("/attendance/students/report", response_model=List[StudentAttendanceReportItemResponse])
async def get_students_attendance_report(
    student_name: str = Query(..., min_length=2),
    school_id: Optional[int] = Query(None),
    class_id: Optional[int] = Query(None),
    academic_year: Optional[int] = Query(None),
    term: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """بحث حضور/غياب طالب بالاسم مع الأيام والتواريخ"""
    purge_expired_archived_submissions(db)
    target_school_id = current_user.school_id
    if school_id and current_user.is_super_admin:
        target_school_id = school_id

    term_settings = ensure_school_term_settings(db, target_school_id)
    selected_year = academic_year if academic_year is not None else term_settings.current_academic_year
    selected_term = normalize_term_or_raise(term) if term else term_settings.current_term

    students_query = db.query(Student).join(Class).filter(
        Class.school_id == target_school_id,
        Student.is_active == True,
        Student.full_name.ilike(f"%{student_name.strip()}%")
    )

    if class_id:
        students_query = students_query.filter(Student.class_id == class_id)

    matched_students = students_query.order_by(Student.full_name).all()

    report_items = []
    for student in matched_students:
        records_query = db.query(AttendanceRecord).join(AttendanceSubmission).join(Class).filter(
            AttendanceRecord.student_id == student.id,
            Class.school_id == target_school_id,
            AttendanceSubmission.deleted_at.is_(None),
            AttendanceSubmission.academic_year == selected_year,
            AttendanceSubmission.term == selected_term
        )

        if class_id:
            records_query = records_query.filter(AttendanceSubmission.class_id == class_id)

        records = records_query.order_by(AttendanceRecord.date.asc(), AttendanceRecord.created_at.asc()).all()

        by_date_status: Dict[date, str] = {}
        for record in records:
            by_date_status[record.date] = record.status.value

        details = [
            StudentAttendanceDateResponse(date=day, status=status)
            for day, status in sorted(by_date_status.items(), key=lambda item: item[0])
        ]

        present_days = sum(1 for item in details if item.status == AttendanceStatusEnum.PRESENT.value)
        absent_days = sum(1 for item in details if item.status == AttendanceStatusEnum.ABSENT.value)

        report_items.append(StudentAttendanceReportItemResponse(
            student_id=student.id,
            student_name=student.full_name,
            class_name=student.class_obj.name if student.class_obj else "-",
            present_days=present_days,
            absent_days=absent_days,
            details=details
        ))

    return report_items


@app.put("/attendance/submissions/{submission_id}/records")
async def update_submission_records(
    submission_id: int,
    request: AttendanceSubmissionUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """تعديل حالات طلاب سجل حضور كامل (إدارة فقط)"""
    purge_expired_archived_submissions(db)
    ensure_admin_user(current_user)

    submission = db.query(AttendanceSubmission).join(Class).filter(
        AttendanceSubmission.id == submission_id,
        AttendanceSubmission.deleted_at.is_(None)
    ).first()
    if not submission:
        raise HTTPException(status_code=404, detail="السجل غير موجود")

    if not current_user.is_super_admin and submission.class_obj.school_id != current_user.school_id:
        raise HTTPException(status_code=403, detail="غير مصرح")

    if not request.records:
        raise HTTPException(status_code=400, detail="لا توجد تعديلات للحفظ")

    submission_record_ids = {record.id for record in submission.records}
    updated_count = 0

    for item in request.records:
        if item.record_id not in submission_record_ids:
            raise HTTPException(status_code=400, detail=f"السجل الفرعي {item.record_id} لا يتبع لهذا التقديم")

        record = db.query(AttendanceRecord).filter(AttendanceRecord.id == item.record_id).first()
        if not record:
            raise HTTPException(status_code=404, detail=f"سجل الحضور {item.record_id} غير موجود")

        new_status = parse_binary_attendance_status(item.status)

        old_status = record.status
        old_notes = record.notes

        record.status = new_status
        record.notes = item.notes
        record.updated_at = datetime.utcnow()

        if old_status != new_status:
            db.add(AuditLog(
                record_id=record.id,
                old_value=old_status.value,
                new_value=new_status.value,
                field_name="status",
                changed_by=current_user.id
            ))

        if old_notes != item.notes:
            db.add(AuditLog(
                record_id=record.id,
                old_value=old_notes,
                new_value=item.notes or "",
                field_name="notes",
                changed_by=current_user.id
            ))

        updated_count += 1

    submission.status = SubmissionStatusEnum.EDITED
    submission.reviewed_at = None
    submission.reviewed_by = None

    db.commit()

    return {
        "message": "تم حفظ تعديلات السجل بنجاح",
        "submission_id": submission_id,
        "updated_records": updated_count
    }

@app.get("/attendance/report", response_model=DailyReportResponse)
async def get_attendance_report(
    class_id: int = Query(...),
    report_date: date = Query(default_factory=date.today),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """الحصول على تقرير الحضور لتاريخ معين"""
    purge_expired_archived_submissions(db)
    
    # Verify class belongs to user's school
    cls = db.query(Class).filter(
        and_(
            Class.id == class_id,
            Class.school_id == current_user.school_id
        )
    ).first()
    
    if not cls:
        raise HTTPException(status_code=404, detail="الصف غير موجود")

    settings = ensure_school_term_settings(db, cls.school_id)
    
    # Get all students in class
    students = db.query(Student).filter(
        and_(
            Student.class_id == class_id,
            Student.is_active == True
        )
    ).all()
    
    # Get attendance records for the date
    records = db.query(AttendanceRecord).join(Student).join(AttendanceSubmission).filter(
        and_(
            Student.class_id == class_id,
            AttendanceRecord.date == report_date,
            AttendanceSubmission.deleted_at.is_(None),
            AttendanceSubmission.academic_year == settings.current_academic_year,
            AttendanceSubmission.term == settings.current_term
        )
    ).all()
    
    # Count statuses
    present_count = sum(1 for r in records if r.status == AttendanceStatusEnum.PRESENT)
    absent_count = sum(1 for r in records if r.status == AttendanceStatusEnum.ABSENT)
    
    total_students = len(students)
    attendance_percentage = (present_count / total_students * 100) if total_students > 0 else 0
    
    return DailyReportResponse(
        class_name=cls.name,
        date=report_date,
        total_students=total_students,
        present_count=present_count,
        absent_count=absent_count,
        attendance_percentage=round(attendance_percentage, 2)
    )

# ===== Admin Routes =====

@app.put("/attendance/records/{record_id}")
async def edit_attendance_record(
    record_id: int,
    status: str = Query(...),
    notes: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """تعديل سجل حضور (للإداريين فقط)"""
    
    # Check if user is admin
    if current_user.role != RoleEnum.ADMIN:
        raise HTTPException(status_code=403, detail="صلاحيات غير كافية")
    
    record = db.query(AttendanceRecord).filter(AttendanceRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="السجل غير موجود")
    
    # Log the change
    old_status = record.status.value
    record.status = AttendanceStatusEnum(status)
    record.notes = notes
    record.updated_at = datetime.utcnow()
    
    audit_log = AuditLog(
        record_id=record_id,
        old_value=old_status,
        new_value=status,
        field_name="status",
        changed_by=current_user.id
    )
    
    db.add(audit_log)
    db.commit()
    db.refresh(record)
    
    return {
        "message": "تم تحديث السجل بنجاح",
        "record_id": record.id,
        "new_status": record.status.value
    }

@app.patch("/attendance/submissions/{submission_id}/approve")
async def approve_submission(
    submission_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """الموافقة على تقديم حضور (للإداريين فقط)"""
    
    if current_user.role != RoleEnum.ADMIN:
        raise HTTPException(status_code=403, detail="صلاحيات غير كافية")
    
    submission = db.query(AttendanceSubmission).filter(
        AttendanceSubmission.id == submission_id
    ).first()
    
    if not submission:
        raise HTTPException(status_code=404, detail="التقديم غير موجود")
    
    submission.status = SubmissionStatusEnum.APPROVED
    submission.reviewed_at = datetime.utcnow()
    submission.reviewed_by = current_user.id
    
    db.commit()
    
    return {
        "message": "تم الموافقة على التقديم",
        "submission_id": submission.id,
        "status": submission.status.value
    }


@app.get("/schools/{school_id}/term-settings")
async def get_school_term_settings(
    school_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user.is_super_admin and current_user.school_id != school_id:
        raise HTTPException(status_code=403, detail="غير مصرح")

    settings = ensure_school_term_settings(db, school_id)
    return {
        "school_id": school_id,
        "academic_year": settings.current_academic_year,
        "term": settings.current_term
    }


@app.put("/schools/{school_id}/term-settings")
async def update_school_term_settings(
    school_id: int,
    request: SchoolTermSettingsUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user.is_super_admin and current_user.school_id != school_id:
        raise HTTPException(status_code=403, detail="غير مصرح")
    ensure_principal_user(current_user)

    normalized_term = normalize_term_or_raise(request.term)
    if request.academic_year < 2000 or request.academic_year > 2100:
        raise HTTPException(status_code=400, detail="السنة الدراسية غير صحيحة")

    settings = ensure_school_term_settings(db, school_id)
    settings.current_academic_year = request.academic_year
    settings.current_term = normalized_term
    settings.updated_by = current_user.id
    settings.updated_at = datetime.utcnow()

    db.commit()
    return {
        "message": "تم تحديث السنة والفصل الحاليين بنجاح",
        "school_id": school_id,
        "academic_year": settings.current_academic_year,
        "term": settings.current_term
    }


@app.post("/admin/attendance/archive-term")
async def archive_term_submissions(
    request: AttendanceArchiveTermRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    ensure_principal_user(current_user)
    purge_expired_archived_submissions(db)

    target_school_id = request.school_id if (request.school_id and current_user.is_super_admin) else current_user.school_id
    normalized_term = normalize_term_or_raise(request.term)

    submissions = db.query(AttendanceSubmission).join(Class).filter(
        Class.school_id == target_school_id,
        AttendanceSubmission.academic_year == request.academic_year,
        AttendanceSubmission.term == normalized_term,
        AttendanceSubmission.deleted_at.is_(None)
    ).all()

    if not submissions:
        return {
            "message": "لا توجد سجلات مطابقة للأرشفة",
            "school_id": target_school_id,
            "academic_year": request.academic_year,
            "term": normalized_term,
            "archived_submissions": 0
        }

    now = datetime.utcnow()
    purge_at = now + timedelta(days=ARCHIVE_RETENTION_DAYS)
    for sub in submissions:
        sub.deleted_at = now
        sub.purge_at = purge_at

    db.commit()
    return {
        "message": f"تم نقل السجلات إلى المحذوف المؤقت لمدة {ARCHIVE_RETENTION_DAYS} أيام",
        "school_id": target_school_id,
        "academic_year": request.academic_year,
        "term": normalized_term,
        "archived_submissions": len(submissions),
        "retention_days": ARCHIVE_RETENTION_DAYS,
        "purge_at": purge_at.isoformat()
    }


@app.post("/admin/attendance/archive-term/restore")
async def restore_archived_term_submissions(
    request: AttendanceArchiveTermRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    ensure_principal_user(current_user)
    purge_expired_archived_submissions(db)

    target_school_id = request.school_id if (request.school_id and current_user.is_super_admin) else current_user.school_id
    normalized_term = normalize_term_or_raise(request.term)

    submissions = db.query(AttendanceSubmission).join(Class).filter(
        Class.school_id == target_school_id,
        AttendanceSubmission.academic_year == request.academic_year,
        AttendanceSubmission.term == normalized_term,
        AttendanceSubmission.deleted_at.isnot(None),
        AttendanceSubmission.purge_at.isnot(None),
        AttendanceSubmission.purge_at > datetime.utcnow()
    ).all()

    for sub in submissions:
        sub.deleted_at = None
        sub.purge_at = None

    db.commit()
    return {
        "message": "تم استرجاع السجلات بنجاح",
        "school_id": target_school_id,
        "academic_year": request.academic_year,
        "term": normalized_term,
        "restored_submissions": len(submissions)
    }


@app.get("/admin/attendance/terms")
async def list_school_terms_overview(
    school_id: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    ensure_principal_user(current_user)
    purge_expired_archived_submissions(db)

    target_school_id = school_id if (school_id and current_user.is_super_admin) else current_user.school_id

    submissions = db.query(AttendanceSubmission).join(Class).filter(
        Class.school_id == target_school_id
    ).order_by(AttendanceSubmission.academic_year.desc(), AttendanceSubmission.created_at.desc()).all()

    overview_map: Dict[str, Dict[str, Any]] = {}
    for sub in submissions:
        key = f"{sub.academic_year}:{sub.term}"
        if key not in overview_map:
            overview_map[key] = {
                "academic_year": sub.academic_year,
                "term": sub.term,
                "total_submissions": 0,
                "active_submissions": 0,
                "archived_submissions": 0,
                "last_submission_date": None,
            }

        item = overview_map[key]
        item["total_submissions"] += 1
        if sub.deleted_at is None:
            item["active_submissions"] += 1
        else:
            item["archived_submissions"] += 1

        if not item["last_submission_date"] or sub.date > datetime.fromisoformat(item["last_submission_date"]).date():
            item["last_submission_date"] = sub.date.isoformat()

    items = sorted(
        overview_map.values(),
        key=lambda row: (row["academic_year"], row["term"]),
        reverse=True,
    )

    return {
        "school_id": target_school_id,
        "terms": items,
    }


@app.get("/admin/attendance/terms/{academic_year}/{term}/details")
async def get_term_details(
    academic_year: int,
    term: str,
    school_id: Optional[int] = Query(None),
    include_archived: bool = Query(True),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    ensure_principal_user(current_user)
    purge_expired_archived_submissions(db)

    normalized_term = normalize_term_or_raise(term)
    target_school_id = school_id if (school_id and current_user.is_super_admin) else current_user.school_id

    query = db.query(AttendanceSubmission).join(Class).filter(
        Class.school_id == target_school_id,
        AttendanceSubmission.academic_year == academic_year,
        AttendanceSubmission.term == normalized_term,
    )
    if not include_archived:
        query = query.filter(AttendanceSubmission.deleted_at.is_(None))

    submissions = query.order_by(AttendanceSubmission.date.desc()).all()

    submission_rows: List[Dict[str, Any]] = []
    total_present = 0
    total_absent = 0
    unique_dates = set()

    for sub in submissions:
        present_count = sum(1 for r in sub.records if r.status == AttendanceStatusEnum.PRESENT)
        absent_count = sum(1 for r in sub.records if r.status == AttendanceStatusEnum.ABSENT)
        total_present += present_count
        total_absent += absent_count
        unique_dates.add(sub.date.isoformat())

        submission_rows.append({
            "submission_id": sub.id,
            "date": sub.date.isoformat(),
            "class_id": sub.class_id,
            "class_name": sub.class_obj.name if sub.class_obj else "-",
            "teacher_name": sub.submitted_by_user.full_name if sub.submitted_by_user else "-",
            "total_students": len(sub.records),
            "present_count": present_count,
            "absent_count": absent_count,
            "is_archived": sub.deleted_at is not None,
            "deleted_at": sub.deleted_at.isoformat() if sub.deleted_at else None,
            "records": [
                {
                    "record_id": r.id,
                    "student_id": r.student_id,
                    "student_name": r.student.full_name if r.student else "-",
                    "status": r.status.value,
                    "date": r.date.isoformat(),
                }
                for r in sub.records
            ]
        })

    return {
        "school_id": target_school_id,
        "academic_year": academic_year,
        "term": normalized_term,
        "summary": {
            "submissions_count": len(submissions),
            "dates_count": len(unique_dates),
            "present_total": total_present,
            "absent_total": total_absent,
        },
        "submissions": submission_rows,
    }
# ===== Export Excel =====
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

@app.get("/admin/attendance/export/excel")
async def export_attendance_excel(
    academic_year: int = Query(...),
    term: str = Query(...),
    school_id: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """تصدير سجلات الحضور والغياب كملف Excel"""
    ensure_principal_user(current_user)
    normalized_term = normalize_term_or_raise(term)
    target_school_id = school_id if (school_id and current_user.is_super_admin) else current_user.school_id

    submissions = db.query(AttendanceSubmission).join(Class).filter(
        Class.school_id == target_school_id,
        AttendanceSubmission.academic_year == academic_year,
        AttendanceSubmission.term == normalized_term,
        AttendanceSubmission.deleted_at.is_(None)
    ).order_by(AttendanceSubmission.date).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سجل الحضور والغياب"
    ws.sheet_view.rightToLeft = True

    # Header style
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    headers = ["التاريخ", "الصف", "المعلم", "اسم الطالب", "الحالة"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    current_row = 2
    for sub in submissions:
        for record in sub.records:
            ws.cell(row=current_row, column=1, value=sub.date.strftime("%Y-%m-%d"))
            ws.cell(row=current_row, column=2, value=sub.class_obj.name if sub.class_obj else "-")
            ws.cell(row=current_row, column=3, value=sub.submitted_by_user.full_name if sub.submitted_by_user else "-")
            ws.cell(row=current_row, column=4, value=record.student.full_name if record.student else "-")
            status = "حاضر ✅" if record.status == AttendanceStatusEnum.PRESENT else "غائب ❌"
            ws.cell(row=current_row, column=5, value=status)
            current_row += 1

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"attendance_{academic_year}_{normalized_term}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ===== Error Handlers =====
@app.exception_handler(HTTPException)
async def http_exception_handler(request, exc):
    return JSONResponse(
        status_code=exc.status_code,
        content={
            "detail": exc.detail,
            "error": exc.detail,
            "status_code": exc.status_code,
            "timestamp": datetime.utcnow().isoformat()
        }
    )

# ===== Static Files Setup =====
# Serve frontend files AFTER all API routes are defined
frontend_path = os.path.join(parent_dir, "frontend")
if os.path.exists(frontend_path):
    app.mount("/", StaticFiles(directory=frontend_path, html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)